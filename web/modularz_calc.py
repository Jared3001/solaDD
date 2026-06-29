"""
modularz_calc.py — server-side calculation engine for the LIHTC v28 workbook.

WHY THIS EXISTS
---------------
The v28 LIHTC model (`models/LIHTC_Model_v28.xlsm`) uses XLOOKUP, LET/LAMBDA,
UNIQUE, Solver and VBA. HyperFormula (the in-browser engine the rest of
ModularZ uses) implements none of those, so the model cannot be computed
client-side. Instead we patch the input cells, recalculate with **LibreOffice
headless**, and read the outputs back. LibreOffice 24.8+ evaluates the model
faithfully (verified to ~6 sig figs against Excel's cached values) once one
quirk is handled — see STRIP below.

THE LIBREOFFICE QUIRKS (both handled here)
------------------------------------------
1. RECALC ON LOAD. By default LibreOffice trusts the cached values of a foreign
   workbook. We force a full recalc with a throwaway user profile whose
   registrymodifications.xcu sets OOXML/ODF RecalcMode = 0 (always).
2. LAMBDA/LET ARTIFACT NAMES. Excel emits hidden global defined names for
   LAMBDA/LET parameters (`_xlpm.*`, `_xleta.*`, value `#NAME?`). On import
   LibreOffice resolves the LET parameter to that broken global name instead of
   binding it locally, so every LET/LAMBDA cell returns `#NAME?` and poisons the
   whole returns chain. Stripping those defined names (they are pure artifacts)
   fixes it completely. We do this in-memory per call.

Both input-patching and the workbook.xml edit are done with surgical XML
rewrites inside the zip — we never round-trip the file through openpyxl (which
would risk mangling the LAMBDA/dynamic-array plumbing). openpyxl is used only to
READ the recalculated output (data_only).

The input/output cell addresses are resolved from the workbook's own
"Cell Mapping" sheet, so when the model author ships v29 the tool re-syncs as
long as that sheet stays current.
"""
import os
import re
import shutil
import subprocess
import tempfile
import zipfile

import threading

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(HERE, "models", "LIHTC_Model_v30.xlsm")

# Pro_Forma is the main sheet; in v28 it is stored as sheet3.xml. Resolved
# dynamically (below) so a re-ordered workbook still works.
PRO_FORMA = "Pro_Forma"

_PROFILE_DIR = None  # cached LibreOffice user-profile path (for the single recalc() path)
_RECALC_SEM  = threading.Semaphore(2)  # max concurrent LibreOffice processes (memory cap)


# --------------------------------------------------------------------------
# LibreOffice discovery + profile
# --------------------------------------------------------------------------
def _soffice_bin():
    """Locate the LibreOffice binary across local (mac) and Railway (nix)."""
    env = os.environ.get("LIBREOFFICE_BIN")
    if env and os.path.exists(env):
        return env
    for cand in ("soffice", "libreoffice"):
        p = shutil.which(cand)
        if p:
            return p
    mac = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(mac):
        return mac
    raise RuntimeError(
        "LibreOffice not found. Set LIBREOFFICE_BIN or install 'libreoffice'."
    )


def _make_profile():
    """Create a throwaway LibreOffice user profile that forces full recalc.
    Returns the directory path. Caller is responsible for cleanup."""
    d = tempfile.mkdtemp(prefix="modularz_loprofile_")
    user = os.path.join(d, "user")
    os.makedirs(user, exist_ok=True)
    with open(os.path.join(user, "registrymodifications.xcu"), "w") as f:
        f.write(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
            'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            "</oor:items>\n"
        )
    return d


def _profile_uri():
    """Shared LibreOffice user profile for the single-threaded recalc() path."""
    global _PROFILE_DIR
    if _PROFILE_DIR and os.path.isdir(_PROFILE_DIR):
        return "file://" + _PROFILE_DIR
    _PROFILE_DIR = _make_profile()
    return "file://" + _PROFILE_DIR


# --------------------------------------------------------------------------
# Workbook plumbing
# --------------------------------------------------------------------------
_DEFNAME_ARTIFACT = re.compile(
    r'<definedName name="_(?:xlpm|xleta)\.[^"]*"[^>]*>.*?</definedName>', re.S
)


def _sheet_target(zin, sheet_name):
    """Resolve 'Pro_Forma' -> 'xl/worksheets/sheetN.xml' via workbook rels."""
    wb = zin.read("xl/workbook.xml").decode("utf-8")
    m = re.search(r'<sheet name="%s"[^>]*r:id="([^"]+)"' % re.escape(sheet_name), wb)
    if not m:
        raise ValueError(f"sheet {sheet_name!r} not found")
    rid = m.group(1)
    rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rm = re.search(r'Id="%s"[^>]*Target="([^"]+)"' % re.escape(rid), rels)
    if not rm:
        raise ValueError(f"rel {rid} not found")
    tgt = rm.group(1)
    if not tgt.startswith("xl/"):
        tgt = "xl/" + tgt.lstrip("/")
    return tgt


def _set_cell_xml(xml, addr, value, is_text):
    """Surgically set a cell's value in worksheet XML, dropping any formula."""
    if is_text:
        safe = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        newc = f'<c r="{addr}" t="inlineStr"><is><t>{safe}</t></is></c>'
    else:
        newc = f'<c r="{addr}"><v>{value}</v></c>'
    pat = re.compile(r'<c r="%s"(?:[^>]*?)(?:/>|>.*?</c>)' % re.escape(addr), re.S)
    if not pat.search(xml):
        raise ValueError(f"cell {addr} not present in sheet XML (cannot patch)")
    return pat.sub(newc, xml, count=1)


# --------------------------------------------------------------------------
# Output map (parsed from the workbook's own "Cell Mapping" sheet)
# --------------------------------------------------------------------------
_OUTPUT_MAP = None
_ADDR_RE = re.compile(r"ADDRESS\(ROW\(Pro_Forma!([A-Z]+\d+)\)", re.I)


def output_map():
    """label -> Pro_Forma cell, parsed from the 'Cell Mapping' sheet.

    Only single-cell ADDRESS(ROW(...)) entries are returned (ranges/headers and
    the one stale #REF! mapping row are skipped)."""
    global _OUTPUT_MAP
    if _OUTPUT_MAP is not None:
        return _OUTPUT_MAP
    wb = openpyxl.load_workbook(MODEL_PATH, data_only=False, keep_vba=True)
    cm = wb["Cell Mapping"]
    out = {}
    for row in cm.iter_rows(min_col=2, max_col=3):  # B = label, C = ADDRESS formula
        label = row[0].value
        formula = row[1].value
        if not label or not isinstance(formula, str):
            continue
        m = _ADDR_RE.search(formula)
        if m:
            out[str(label).strip()] = m.group(1)
    _OUTPUT_MAP = out
    return out


# Curated subset surfaced to the UI as headline numbers.
HEADLINE = {
    "Equity Required": "C24",
    "GP IRR (15-yr)": "C25",
    "GP MOIC": "C26",
    "GP Net Profit": "C27",
    "Total IRR": "C28",
    "Tiebreak Score": "C30",
    "Total Uses": "D55",
    "Tax Credit Equity": "D80",
    "Permanent Loan": "D75",
}

# Cells read per scenario for the PDF comparison table (HEADLINE + unit count + NRSF).
# Plus dev-fee tranche cells for the deal-tracker export:
#   C47 = Developer Fee (total)            D97 = Capitalized Dev Fee (= C47 - D71)
#   D71 = Deferred Developer Fee           C24 = Sponsor Funds Needed (in HEADLINE)
#   C30 = Tiebreak Score (in HEADLINE)
# The tracker's "Capitalized Dev Fee 30%/70%" columns are a fixed partition of D97.
SCENARIO_READ = list(HEADLINE.values()) + ["C14", "C17", "C47", "D97", "D71"]


# Friendly input name -> (Pro_Forma cell, is_text). The true editable inputs
# (constants), distinct from the Cell Mapping's display cells which are derived.
INPUT_CELLS = {
    "project_name":        ("B2",  True),
    "county":              ("C3",  True),
    "lot_sf":              ("C12", False),   # site area; drives acres -> units
    "residential_stories": ("C15", False),
    "podium_levels":       ("C18", False),
    "rent_growth":         ("C19", False),
    "expense_growth":      ("C20", False),
    "bond_test_limit":     ("C21", False),
    "land_cost":           ("S16", False),   # acquisition price
    "tax_exempt":          ("S26", True),    # "Yes"/"No"
    "interest_rate":       ("S27", False),
    "dscr":                ("S28", False),
    "amortization":        ("S29", False),
    "construction_type":   ("A36", True),    # "Modular"/"Stick"
    "gap_financing":       ("D59", True),    # Ground Lessor/Soft Debt/State Credits/B-Bond/None
    # Unit mix as a fraction of total units (1B = 1 - 0B - 2B - 3B, derived in-model).
    "unit_mix_0b":         ("I3",  False),
    "unit_mix_2b":         ("I5",  False),
    "unit_mix_3b":         ("I6",  False),
}

# Options for the gap-financing dropdown (from the workbook's data validation on D59).
GAP_FINANCING_OPTIONS = ["None", "Ground Lessor", "Soft Debt", "State Credits", "B-Bond"]


def _scenario_inputs(dd, scenario, deal_name=None):
    """Return (inputs_num, inputs_txt) for one scenario.

    Shared between build_for_scenario (Excel download) and recalc_isolated
    (PDF output computation) so both always use identical inputs."""
    import uw_logic
    base, _meta = uw_logic.base_cells(dd)
    cells = dict(base)
    try:
        cells.update(uw_logic.method_cells(scenario["constr"]))
    except Exception:  # noqa: BLE001
        pass

    derived = _v28_formula_cells()
    num, txt = {}, {}
    for key, val in cells.items():
        sheet, ref = key if isinstance(key, tuple) else (PRO_FORMA, key)
        if sheet != PRO_FORMA or val is None or ref in derived:
            continue
        (txt if isinstance(val, str) else num)[ref] = val

    if "S16" not in num:
        lot = _parse_num(dd.get("land_sf"))
        if lot:
            num["S16"] = round(lot * 150)

    # Scenario overrides bypass the formula filter (intentional).
    num.update({
        "C15": int(scenario["stories"]),
        "C18": int(scenario["podium"]),
        "I3":  0.0,
        "I5":  float(scenario["sh2B"]),
        "I6":  float(scenario["sh3B"]),
    })
    sf_key = scenario["constr"]
    if sf_key == "Stick" and scenario.get("lf") == "Yes":
        sf_key = "Stick_LF"
    num.update(SCENARIO_SF.get(sf_key, {}))
    txt["C11"] = scenario["lf"]

    if deal_name:
        txt["B2"] = str(deal_name)

    return num, txt


def _parse_num(s):
    """'$15,000,000' / '20,000 SF' / '1.2M' -> float, or None if nothing parses."""
    if s is None:
        return None
    t = str(s).strip().lower().replace(",", "").replace("$", "")
    import re as _re
    m = _re.search(r"(-?\d*\.?\d+)\s*(k|m|mm|million|thousand)?", t)
    if not m:
        return None
    n = float(m.group(1))
    u = m.group(2) or ""
    if u in ("k", "thousand"):
        n *= 1e3
    elif u in ("m", "mm", "million"):
        n *= 1e6
    return n


def om_fields_to_inputs(fields):
    """Map OM-extracted {field_id: value} onto the v28 friendly inputs, parsing
    numbers. Implements Cazalis's land default chain: OM acquisition price first.
    Returns {friendly_name: value} for whatever the OM provided."""
    out = {}
    addr = (fields.get("address") or "").strip()
    if addr:
        out["project_name"] = addr
    price = _parse_num(fields.get("acquisition_price"))
    if price and price > 0:
        out["land_cost"] = round(price)          # OM price wins over the $150/SF default
    lot = _parse_num(fields.get("land_sf"))
    if lot and lot > 0:
        out["lot_sf"] = round(lot)
    # County: match the OM's county/jurisdiction to a workbook county (drives rents).
    cand = (fields.get("county") or fields.get("city_jurisdiction") or "").strip()
    if cand:
        low = cand.lower()
        for c in counties():
            if c.lower() == low or c.lower() in low:
                out["county"] = c
                break
    return out


def research_address(address):
    """Use Gemini to resolve a bare address into v28 inputs: county (drives the
    AMI rent lookups) + a rough land $/SF estimate. Returns {inputs, land_psf, city}."""
    import os
    import json as _json
    import requests
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not key:
        raise RuntimeError("Address research needs GEMINI_API_KEY set on the server.")
    cs = counties()
    prompt = (
        f'For the property at "{address}", return JSON with: '
        f'"county" (the California county it is in — choose EXACTLY from this list when applicable: '
        f'{", ".join(cs)}), "city", and "land_psf" (a rough land value in USD per square foot of LOT '
        f'for that submarket, as a number). Use null for anything you cannot determine.'
    )
    schema = {
        "type": "object",
        "properties": {
            "county": {"type": "string"},
            "city": {"type": "string"},
            "land_psf": {"type": "number"},
        },
    }
    body = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.1, "responseMimeType": "application/json",
                             "responseSchema": schema},
    }
    r = requests.post(
        "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key=" + key,
        json=body, timeout=30)
    r.raise_for_status()
    txt = r.json()["candidates"][0]["content"]["parts"][0]["text"]
    obj = _json.loads(txt)
    inputs = {"project_name": address}
    craw = (obj.get("county") or "").strip().lower()
    for c in cs:
        if c.lower() == craw or (craw and c.lower() in craw):
            inputs["county"] = c
            break
    return {"inputs": inputs, "land_psf": obj.get("land_psf"), "city": obj.get("city")}


def split_friendly(inputs):
    """{friendly_name: value} -> (inputs_num, inputs_txt) keyed by cell."""
    num, txt = {}, {}
    for name, val in (inputs or {}).items():
        spec = INPUT_CELLS.get(name)
        if not spec:
            continue
        cell, is_text = spec
        (txt if is_text else num)[cell] = val
    return num, txt


# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------
def recalc(inputs_num=None, inputs_txt=None, read=None, timeout=120):
    """Patch Pro_Forma input cells, recalc with LibreOffice, read outputs.

    inputs_num / inputs_txt : {cell_addr: value} on the Pro_Forma sheet.
    read : iterable of cell addresses to return; defaults to HEADLINE values.
    Returns {cell_addr: value}.
    """
    inputs_num = inputs_num or {}
    inputs_txt = inputs_txt or {}
    read = list(read) if read else list(HEADLINE.values())

    work = tempfile.mkdtemp(prefix="modularz_calc_")
    try:
        clean = os.path.join(work, "model.xlsm")
        with zipfile.ZipFile(MODEL_PATH, "r") as zin:
            pf_target = _sheet_target(zin, PRO_FORMA)
            wbxml = _DEFNAME_ARTIFACT.sub("", zin.read("xl/workbook.xml").decode("utf-8"))
            pfxml = zin.read(pf_target).decode("utf-8")
            for a, v in inputs_num.items():
                pfxml = _set_cell_xml(pfxml, a, v, False)
            for a, v in inputs_txt.items():
                pfxml = _set_cell_xml(pfxml, a, v, True)
            with zipfile.ZipFile(clean, "w", zipfile.ZIP_DEFLATED) as zout:
                for it in zin.infolist():
                    data = zin.read(it.filename)
                    if it.filename == "xl/workbook.xml":
                        data = wbxml.encode("utf-8")
                    elif it.filename == pf_target:
                        data = pfxml.encode("utf-8")
                    zout.writestr(it, data)

        outdir = os.path.join(work, "out")
        proc = subprocess.run(
            [_soffice_bin(), "--headless", "--nologo", "--nofirststartwizard",
             f"-env:UserInstallation={_profile_uri()}",
             "--convert-to", "xlsx", "--outdir", outdir, clean],
            capture_output=True, timeout=timeout,
        )
        out_xlsx = os.path.join(outdir, "model.xlsx")
        if not os.path.exists(out_xlsx):
            raise RuntimeError(
                "LibreOffice recalc produced no output. stderr=%r"
                % proc.stderr.decode("utf-8", "replace")[:500]
            )
        ws = openpyxl.load_workbook(out_xlsx, data_only=True)[PRO_FORMA]
        return {c: ws[c].value for c in read}
    finally:
        shutil.rmtree(work, ignore_errors=True)


def recalc_isolated(inputs_num=None, inputs_txt=None, read=None, timeout=120):
    """Like recalc() but creates its own throwaway LibreOffice profile per call,
    making it safe to call from multiple threads simultaneously.
    Concurrency is capped at 2 via _RECALC_SEM to avoid OOM on Railway."""
    inputs_num = inputs_num or {}
    inputs_txt = inputs_txt or {}
    read = list(read) if read else list(HEADLINE.values())

    with _RECALC_SEM:
        profile_dir = _make_profile()
        work = tempfile.mkdtemp(prefix="modularz_calc_")
        try:
            clean = os.path.join(work, "model.xlsm")
            with zipfile.ZipFile(MODEL_PATH, "r") as zin:
                pf_target = _sheet_target(zin, PRO_FORMA)
                wbxml = _DEFNAME_ARTIFACT.sub("", zin.read("xl/workbook.xml").decode("utf-8"))
                pfxml = zin.read(pf_target).decode("utf-8")
                for a, v in inputs_num.items():
                    pfxml = _set_cell_xml(pfxml, a, v, False)
                for a, v in inputs_txt.items():
                    pfxml = _set_cell_xml(pfxml, a, v, True)
                with zipfile.ZipFile(clean, "w", zipfile.ZIP_DEFLATED) as zout:
                    for it in zin.infolist():
                        data = zin.read(it.filename)
                        if it.filename == "xl/workbook.xml":
                            data = wbxml.encode("utf-8")
                        elif it.filename == pf_target:
                            data = pfxml.encode("utf-8")
                        zout.writestr(it, data)

            outdir = os.path.join(work, "out")
            os.makedirs(outdir, exist_ok=True)
            proc = subprocess.run(
                [_soffice_bin(), "--headless", "--nologo", "--nofirststartwizard",
                 f"-env:UserInstallation=file://{profile_dir}",
                 "--convert-to", "xlsx", "--outdir", outdir, clean],
                capture_output=True, timeout=timeout,
            )
            out_xlsx = os.path.join(outdir, "model.xlsx")
            if not os.path.exists(out_xlsx):
                raise RuntimeError(
                    "LibreOffice recalc produced no output. stderr=%r"
                    % proc.stderr.decode("utf-8", "replace")[:500]
                )
            ws = openpyxl.load_workbook(out_xlsx, data_only=True)[PRO_FORMA]
            return {c: ws[c].value for c in read}
        finally:
            shutil.rmtree(work, ignore_errors=True)
            shutil.rmtree(profile_dir, ignore_errors=True)


def _force_full_recalc(wbxml):
    """Set fullCalcOnLoad=1 so Excel recalculates the patched workbook on open."""
    m = re.search(r"<calcPr[^>]*?/?>", wbxml)
    if not m:
        return wbxml
    tag = m.group(0)
    if "fullCalcOnLoad" in tag:
        new = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', tag)
    elif tag.endswith("/>"):
        new = tag[:-2] + ' fullCalcOnLoad="1"/>'
    else:
        new = tag[:-1] + ' fullCalcOnLoad="1">'
    return wbxml.replace(tag, new, 1)


# Orphan scratch sheets in the source model that carry stale #VALUE/#DIV errors
# (unreferenced by the proforma) — hidden in the download so they don't read as
# "broken" when the file is opened in Excel.
_SUPPRESS_SHEETS = ("Sheet1",)


def _hide_sheet(wbxml, name):
    """Set state="hidden" on a <sheet> entry in workbook.xml (if present)."""
    pat = re.compile(r'(<sheet name="%s"[^>]*?)(\s*/>)' % re.escape(name))

    def repl(m):
        head = m.group(1)
        if "state=" in head:
            return m.group(0)
        return head + ' state="hidden"/>'

    return pat.sub(repl, wbxml, count=1)


def build_download(inputs_num=None, inputs_txt=None, tolerant=False):
    """Return the .xlsm with the input cells patched and full-recalc-on-open set.
    Unlike recalc(), this keeps the workbook PRISTINE (macros, data validations,
    and the LAMBDA/LET defined names intact) so Excel computes it natively.
    tolerant=True skips any patch cell that isn't present (for externally-supplied
    cell maps that may include cells this model doesn't have)."""
    inputs_num = inputs_num or {}
    inputs_txt = inputs_txt or {}
    import io
    buf = io.BytesIO()

    def _patch(xml, addr, val, is_text):
        try:
            return _set_cell_xml(xml, addr, val, is_text)
        except ValueError:
            if tolerant:
                return xml
            raise

    with zipfile.ZipFile(MODEL_PATH, "r") as zin:
        pf_target = _sheet_target(zin, PRO_FORMA)
        pfxml = zin.read(pf_target).decode("utf-8")
        for a, v in inputs_num.items():
            pfxml = _patch(pfxml, a, v, False)
        for a, v in inputs_txt.items():
            pfxml = _patch(pfxml, a, v, True)
        wbxml = zin.read("xl/workbook.xml").decode("utf-8")
        # Strip the junk LAMBDA/LET artifact names (cached #NAME?) — they break the
        # LET/LAMBDA chain when Excel RECALCULATES (every Sources/Tax-Credit cell
        # downstream of the hard-cost LET goes #VALUE). Removing them lets the LET
        # bind its parameter locally and recalc clean. (Same fix as the live engine.)
        wbxml = _DEFNAME_ARTIFACT.sub("", wbxml)
        wbxml = _force_full_recalc(wbxml)
        for s in _SUPPRESS_SHEETS:
            wbxml = _hide_sheet(wbxml, s)
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for it in zin.infolist():
                data = zin.read(it.filename)
                if it.filename == pf_target:
                    data = pfxml.encode("utf-8")
                elif it.filename == "xl/workbook.xml":
                    data = wbxml.encode("utf-8")
                zout.writestr(it, data)
    return buf.getvalue()


# Per-unit SF overrides for scenario generation.
# Stick SF splits by Large Family: LF uses 475/735/945, non-LF uses 450/700/900.
SCENARIO_SF = {
    "Modular":   {"L4": 497, "L5": 804, "L6": 994},
    "Stick":     {"L4": 450, "L5": 700, "L6": 900},
    "Stick_LF":  {"L4": 475, "L5": 735, "L6": 945},
}

_V28_FORMULA_CELLS = None


def _v28_formula_cells():
    """Pro_Forma cells that are FORMULAS in v28 — never overwrite these from the DD
    map (e.g. NRSF C17 is derived in v28, per Cazalis's 'NRSF shouldn't be a plug')."""
    global _V28_FORMULA_CELLS
    if _V28_FORMULA_CELLS is None:
        pf = openpyxl.load_workbook(MODEL_PATH, data_only=False, keep_vba=True)[PRO_FORMA]
        _V28_FORMULA_CELLS = {
            c.coordinate for row in pf.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        }
    return _V28_FORMULA_CELLS


def build_from_dd(dd, method="Modular", deal_name=None):
    """Build the v28 LIHTC model (.xlsm) from a DD facts dict, reusing the existing
    DD→Pro_Forma cell mapping (uw_logic.base_cells) + the construction method.
    deal_name (if given) sets the project name (Pro_Forma!B2). Returns .xlsm bytes."""
    import uw_logic  # build/sources — same module the legacy exporter uses
    base, _meta = uw_logic.base_cells(dd)
    cells = dict(base)
    try:
        cells.update(uw_logic.method_cells(method))
    except Exception:  # noqa: BLE001
        pass
    derived = _v28_formula_cells()
    num, txt = {}, {}
    for key, val in cells.items():
        sheet, ref = key if isinstance(key, tuple) else (PRO_FORMA, key)
        if sheet != PRO_FORMA or val is None or ref in derived:
            continue
        (txt if isinstance(val, str) else num)[ref] = val
    # v28 defaults (Cazalis): stories 5 when the DD didn't supply one (the legacy
    # mapper defaults to 3); land at $150/SF of lot when there's no OM price.
    if not dd.get("residential_stories"):
        num["C15"] = 5
    if "S16" not in num:
        lot = _parse_num(dd.get("land_sf"))
        if lot:
            num["S16"] = round(lot * 150)
    if deal_name:
        txt["B2"] = str(deal_name)
    return build_download(num, txt, tolerant=True)


def build_for_scenario(dd, scenario, deal_name=None):
    """Build the v28 LIHTC model (.xlsm) for a specific scenario.

    scenario = {
        "name":    str,
        "constr":  "Modular" | "Stick",
        "stories": int,
        "podium":  int,
        "lf":      "Yes" | "No",
        "sh2B":    float,   # share of units that are 2B (0–1)
        "sh3B":    float,   # share of units that are 3B (0–1)
    }
    """
    num, txt = _scenario_inputs(dd, scenario, deal_name)
    return build_download(num, txt, tolerant=True)


def headline(inputs_num=None, inputs_txt=None):
    """Convenience: recalc and return {friendly label: value}."""
    vals = recalc(inputs_num, inputs_txt, read=list(HEADLINE.values()))
    return {label: vals.get(cell) for label, cell in HEADLINE.items()}


def calc_friendly(inputs):
    """Recalc from {friendly_name: value} inputs; return {friendly label: value}."""
    num, txt = split_friendly(inputs)
    return headline(num, txt)


_COUNTIES = None


def counties():
    """CA county names from the workbook's Market Inputs sheet (drives rent XLOOKUPs)."""
    global _COUNTIES
    if _COUNTIES is not None:
        return _COUNTIES
    wb = openpyxl.load_workbook(MODEL_PATH, data_only=True, keep_vba=True)
    mi = wb["Market Inputs"]
    out = []
    for r in range(4, 62):  # B4:B61
        v = mi.cell(r, 2).value
        if v and str(v).strip():
            out.append(str(v).strip())
    _COUNTIES = out
    return out


def selftest():
    """Lightweight environment check — used by the /api/modularz/health route to
    confirm LibreOffice is installed and can recalc the model on the server."""
    import time
    info = {"model_present": os.path.exists(MODEL_PATH)}
    try:
        info["soffice_bin"] = _soffice_bin()
    except Exception as e:  # noqa: BLE001
        info["soffice_bin"] = None
        info["error"] = str(e)
        info["ok"] = False
        return info
    try:
        ver = subprocess.run([info["soffice_bin"], "--version"],
                             capture_output=True, timeout=30)
        info["soffice_version"] = ver.stdout.decode("utf-8", "replace").strip()[:120]
    except Exception as e:  # noqa: BLE001
        info["soffice_version"] = f"(version check failed: {e})"
    try:
        t0 = time.time()
        gp_irr = recalc(read=["C25"]).get("C25")
        info["recalc_ms"] = int((time.time() - t0) * 1000)
        info["recalc_gp_irr"] = gp_irr
        info["ok"] = isinstance(gp_irr, (int, float))
    except Exception as e:  # noqa: BLE001
        info["ok"] = False
        info["error"] = str(e)
    return info


if __name__ == "__main__":
    import json
    print("output_map entries:", len(output_map()))
    print(json.dumps(headline(), indent=2, default=str))
