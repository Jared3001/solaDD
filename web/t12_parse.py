#!/usr/bin/env python3
"""
t12_parse.py — turn a trailing-12-month (T-12) operating statement into the
non-LIHTC engine's OpEx factors (PUPM = per-unit-per-month).

WHY: the ModularZ market engine takes each OpEx line as a per-unit-per-month
factor on `(Z+) OpEx` col D (see NONLIHTC_ENGINE_SPEC.md). A T-12 gives ANNUAL
dollars per line, so PUPM = annual / units / 12. Management is special — the
engine wants it as a % of revenue, so we divide the T-12 management fee by the
T-12 revenue instead.

DESIGN (mirrors build/budget_extract.py): T-12s have no standard layout, so the
parse is LABEL-DRIVEN and position-tolerant. We scan every sheet, find the one
that reads like an operating statement, then for each expense row match its label
to a friendly OpEx name via an alias map and pull the row's ANNUAL figure (a
detected Total column, an embedded total, a 12-month sum, or the rightmost value).
Property taxes are detected and SKIPPED (the engine derives tax as a formula).
Everything matched/unmatched is reported so the analyst can eyeball + override.

Public API:
    parse_t12(source, units, sheet=None) -> {
        "opex": {friendly_name: pupm_factor, ...},     # feeds nonlihtc.opex
        "management_pct": float | None,                # opex_management (fraction)
        "annual": {friendly_name: annual_dollars},     # what we read, pre-conversion
        "revenue": float | None,                       # detected EGI / total income
        "units": int,
        "matched": [{line, friendly, annual, pupm, basis}],
        "unmatched": [{line, annual}],                 # expense-looking rows we skipped
        "notes": [str, ...],
        "sheet": str,
    }
source may be a path (str/Path) or raw .xlsx bytes.
"""
from __future__ import annotations

import io
import re
from pathlib import Path

import openpyxl


# --------------------------------------------------------------------------- #
# label alias map — friendly OpEx name -> regexes (ordered specific -> generic)
# Each engine OpEx line on (Z+) OpEx col D; management is a % of revenue.
# --------------------------------------------------------------------------- #
ALIASES = [
    ("opex_insurance",   [r"\binsurance\b", r"\bliability\b"]),
    ("opex_management",  [r"management fee", r"property management", r"\bmgmt\b",
                          r"\bmanagement\b"]),
    ("opex_electric",    [r"\belectric", r"\bedison\b", r"\bdwp\b.*electric"]),
    ("opex_water_sewer", [r"water\s*/?\s*&?\s*sewer", r"\bwater\b", r"\bsewer\b"]),
    ("opex_gas",         [r"natural gas", r"socal\s*gas", r"\bgas\b"]),
    ("opex_trash",       [r"\btrash\b", r"\brefuse\b", r"\bgarbage\b",
                          r"waste removal", r"\bwaste\b"]),
    ("opex_landscape",   [r"landscap", r"\bgrounds\b", r"garden"]),
    # elevator BEFORE the generic maintenance line so "Elevator Maintenance" wins
    ("opex_elevator",    [r"\belevator", r"\blift\b"]),
    ("opex_mr_turnover", [r"repairs?\s*&?\s*and?\s*maintenance", r"\br\s*&\s*m\b",
                          r"\bmaintenance\b", r"\bturnover\b", r"make[\s-]?ready",
                          r"\bunit turn", r"\brepairs?\b"]),
    ("opex_payroll",     [r"\bpayroll\b", r"\bsalaries\b", r"\bwages\b",
                          r"on[\s-]?site (staff|personnel|payroll)", r"\bsecurity\b",
                          r"\bpersonnel\b", r"site staff"]),
    ("opex_hcid",        [r"\bhcid\b", r"\bscep\b", r"systematic code",
                          r"housing department"]),
    ("opex_legal",       [r"\blegal\b", r"professional fee", r"\baccounting\b",
                          r"\baudit\b"]),
    ("opex_reserves",    [r"replacement reserve", r"\breserves?\b"]),
    # generic / catch-all goes LAST so specific lines win first
    ("opex_misc",        [r"\bmisc", r"miscellaneous", r"general\s*&?\s*and?\s*admin",
                          r"\bg\s*&\s*a\b", r"\boffice\b", r"administrative",
                          r"other (operating|expense)", r"advertising", r"marketing",
                          r"telephone|internet|cable", r"pest control", r"licenses?\b",
                          r"\bdues\b|permits?"]),
]

# revenue lines (for the management %) — most-specific first
REVENUE_ALIASES = [
    r"effective gross income", r"\begi\b",
    r"net rental income", r"total (operating )?(income|revenue)",
    r"gross (potential|scheduled) (rent|income)", r"\bgpr\b", r"\bgross income\b",
]

# property tax — detected and SKIPPED (engine derives tax as a formula)
TAX_ALIASES = [r"propert(y|ies)\s+tax", r"real estate tax", r"\bre tax\b", r"\btaxes?\b"]

# rows that are sub/grand totals or non-operating — never treated as expense lines
SKIP_ALIASES = [
    r"total (operating )?expense", r"\bnoi\b", r"net operating income",
    r"total expense", r"\bsubtotal\b", r"\btotal\b\s*$", r"debt service",
    r"net income", r"cash flow", r"depreciation", r"amortization",
    r"capital (expenditure|improvement)", r"\bcapex\b",
]


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _match_any(label_n: str, patterns) -> bool:
    return any(re.search(p, label_n) for p in patterns)


def _classify(label_n: str):
    """Return the friendly OpEx name for a label, or None. Specific lines win
    because ALIASES is ordered specific -> generic (misc last)."""
    for friendly, pats in ALIASES:
        if _match_any(label_n, pats):
            return friendly
    return None


def _revenue_rank(label_n: str):
    """Index of the first matching REVENUE_ALIAS (lower = more specific), or None.
    Lets us prefer EGI over Gross Potential Rent for the management % denominator,
    regardless of which appears higher in the statement."""
    if _match_any(label_n, [r"expense"]):
        return None
    for i, p in enumerate(REVENUE_ALIASES):
        if re.search(p, label_n):
            return i
    return None


# --------------------------------------------------------------------------- #
# row reading — label + the row's annual figure
# --------------------------------------------------------------------------- #
def _row_cells(ws, r, max_col):
    label, nums = None, []
    for c in range(1, max_col + 1):
        v = ws.cell(r, c).value
        if _is_num(v):
            nums.append((c, float(v)))
        elif label is None and isinstance(v, str) and v.strip():
            label = v.strip()
    return label, nums


def _detect_total_col(ws, max_row, max_col):
    """Find a column whose header reads Total/Annual/YTD/Year — the annual figure."""
    wants = (r"\btotal\b", r"\bannual\b", r"\bytd\b", r"year\s*to\s*date",
             r"\bt-?12\b", r"trailing")
    for r in range(1, min(max_row, 12) + 1):
        for c in range(1, max_col + 1):
            t = _norm(ws.cell(r, c).value)
            if t and _match_any(t, wants):
                return c
    return None


def _annual_from_row(nums, total_col):
    """Pick the annual dollar figure from a row's numeric cells.
    Priority: detected Total column -> embedded total (one value ~= sum of the
    rest) -> 12-month sum -> the rightmost value."""
    if not nums:
        return None, None
    if total_col is not None:
        for c, v in nums:
            if c == total_col and v:
                return v, "total-col"
    vals = [v for _, v in nums]
    s = sum(vals)
    # an embedded total: a single cell ~= the sum of the others (12 months + total)
    if len(vals) >= 3:
        for c, v in nums:
            others = s - v
            if others > 0 and abs(v - others) <= 0.02 * others:
                return v, "embedded-total"
    if len(vals) >= 11:                       # ~monthly series with no total column
        return s, "sum-monthly"
    return sorted(nums)[-1][1], "rightmost"   # single annual figure (rightmost)


def _score_sheet(ws, max_row, max_col):
    """How operating-statement-like a sheet is = count of expense-label hits."""
    hits = 0
    for r in range(1, min(max_row, 200) + 1):
        label, nums = _row_cells(ws, r, min(max_col, 30))
        if label and nums and _classify(_norm(label)):
            hits += 1
    return hits


# --------------------------------------------------------------------------- #
# main parse
# --------------------------------------------------------------------------- #
def parse_t12(source, units, sheet=None) -> dict:
    units = int(units) if units else 0
    if units <= 0:
        raise ValueError("parse_t12 needs a positive unit count for PUPM conversion.")

    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(Path(source), data_only=True)

    # choose the operating-statement sheet
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        best, best_score = wb.worksheets[0], -1
        for w in wb.worksheets:
            mc = min(w.max_column or 1, 40)
            sc = _score_sheet(w, w.max_row or 1, mc)
            name_bonus = 2 if _match_any(_norm(w.title),
                                         [r"t-?12", r"operating", r"income", r"p\s*&\s*l", r"profit"]) else 0
            if sc + name_bonus > best_score:
                best, best_score = w, sc + name_bonus
        ws = best

    max_row = min(ws.max_row or 1, 400)
    max_col = min(ws.max_column or 1, 40)
    total_col = _detect_total_col(ws, max_row, max_col)

    annual: dict[str, float] = {}
    matched, unmatched, notes = [], [], []
    rev_best = None          # (rank, amount) — most-specific revenue line wins
    seen = set()

    for r in range(1, max_row + 1):
        label, nums = _row_cells(ws, r, max_col)
        if not label or not nums:
            continue
        ln = _norm(label)

        # revenue (for the management %) — pick the most-specific line (EGI over
        # GPR), not just the first by row. Revenue rows never fall to unmatched.
        rank = _revenue_rank(ln)
        if rank is not None:
            rev, _ = _annual_from_row(nums, total_col)
            if rev and rev > 0 and (rev_best is None or rank < rev_best[0]):
                rev_best = (rank, rev)
            continue

        if _match_any(ln, SKIP_ALIASES):
            continue
        if _match_any(ln, TAX_ALIASES):
            amt, _ = _annual_from_row(nums, total_col)
            if amt:
                notes.append(f"Skipped property tax line “{label}” (${amt:,.0f}/yr) "
                             "— the engine derives tax as a formula.")
            continue

        friendly = _classify(ln)
        amt, basis = _annual_from_row(nums, total_col)
        if amt is None or amt == 0:
            continue
        if not friendly:
            unmatched.append({"line": label, "annual": round(amt, 2)})
            continue
        if friendly in seen:                 # first (most specific) row wins per line
            notes.append(f"Additional “{label}” (${amt:,.0f}) folded into {friendly}.")
            annual[friendly] = annual.get(friendly, 0) + amt
            continue
        seen.add(friendly)
        annual[friendly] = amt
        matched.append({"line": label, "friendly": friendly,
                        "annual": round(amt, 2), "basis": basis})

    revenue = rev_best[1] if rev_best else None

    # convert annual $ -> PUPM, except management (% of revenue)
    opex: dict[str, float] = {}
    mgmt_pct = None
    for friendly, amt in annual.items():
        if friendly == "opex_management":
            if revenue and revenue > 0:
                mgmt_pct = round(amt / revenue, 4)
                opex[friendly] = mgmt_pct
            else:
                notes.append("Management fee found but no revenue line — left at the "
                             "engine default (5%).")
            continue
        opex[friendly] = round(amt / units / 12, 2)

    for m in matched:
        if m["friendly"] == "opex_management":
            m["pupm"] = mgmt_pct
            m["unit"] = "% of revenue"
        else:
            m["pupm"] = opex.get(m["friendly"])
            m["unit"] = "$/unit/mo"

    if not matched:
        notes.append("No OpEx lines recognized — check the sheet/layout; the model "
                     "will use the v5.0.7 PUPM defaults.")

    return {
        "opex": opex, "management_pct": mgmt_pct, "annual": annual,
        "revenue": revenue, "units": units, "matched": matched,
        "unmatched": unmatched, "notes": notes, "sheet": ws.title,
    }


if __name__ == "__main__":  # pragma: no cover
    import sys, json
    src = sys.argv[1]
    units = int(sys.argv[2]) if len(sys.argv) > 2 else 50
    print(json.dumps(parse_t12(src, units), indent=2, default=str))
