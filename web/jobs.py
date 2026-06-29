#!/usr/bin/env python3
"""
jobs.py — background job orchestration for the web front end.

This is a thin layer over the existing CLI pipeline. It REUSES the exact
reader registries and runner from build/ (no logic is forked here):

  - single-address runs replicate collect.py's two-phase flow (fan readers out
    across a thread pool, then write the workbook once) but report each field as
    its reader finishes, so the UI can stream live progress.
  - assemblage runs call assemblage.assess() unchanged and surface its result.

Jobs run in daemon threads and live in an in-memory store, so the app MUST run
as a single gunicorn worker (see Procfile). Filled workbooks are written to a
temp dir and served by /api/download.
"""
import os
import sys
import json
import uuid
import shutil
import tempfile
import datetime
import threading
import traceback
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

WEB = Path(__file__).resolve().parent                  # web/
ROOT = WEB.parent                                      # repo root (web/ -> root)
BUILD = ROOT / "build"
SOURCES = BUILD / "sources"
for _p in (str(WEB), str(BUILD), str(SOURCES)):        # make web/ + build/ + build/sources/ importable
    if _p not in sys.path:
        sys.path.insert(0, _p)

import yaml
from openpyxl import load_workbook

# Reuse the CLI pipeline verbatim — registries, runner, geocoder, readers.
import collect as _collect            # READERS, ZIMAS_READERS
import assemblage as _assemblage      # assess()
import underwrite as _underwrite      # export() — DD checklist -> Stick + Modular models
from runner import run_reader, apply_outcome
from geocoder import geocode
import zimas
import lacounty
from jurisdiction import _county_basename
import nc

TEMPLATE = ROOT / "template" / "Checklist_BLANK_master.xlsx"

# Filled workbooks + the persistent counter live in DATA_DIR. Defaults to a
# temp dir (resets on a Railway redeploy); point DATA_DIR at a mounted volume
# for a counter that survives deploys.
DATA_DIR = Path(os.environ.get("DATA_DIR") or (Path(tempfile.gettempdir()) / "sola_dd_runs"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
RUN_DIR = DATA_DIR
COUNTER_FILE = DATA_DIR / "counter.json"
INDEX_FILE = DATA_DIR / "jobs_index.json"   # compact recent-runs index — survives redeploys (files already persist in DATA_DIR)
DEVICE_FILE = DATA_DIR / "devices.json"     # per-device usage tally (silent attribution) — survives redeploys

# Time-saved metric — now spans the whole feasibility pipeline (DD -> Underwriting
# -> Comps -> one-pager), not just the DD checklist. Each completed automated run
# of a stage saves the analyst MINUTES_PER_STAGE[stage]; the banner sums them.
# STARTING_CHECKLISTS = DD sites already automated before the app began counting.
STARTING_CHECKLISTS = int(os.environ.get("STARTING_CHECKLISTS", "9"))
MINUTES_PER_CHECKLIST = int(os.environ.get("MINUTES_PER_CHECKLIST", "30"))  # DD (legacy name)

# Map each job kind to a pipeline stage. Kinds not listed don't accrue time.
STAGE_OF_KIND = {
    "single": "dd", "assemblage": "dd",
    "underwrite": "underwriting", "lihtc_scenarios": "underwriting",
    "comps": "comps", "comps_grid": "comps",
    "pdf_summary": "summary",
}
# Minutes of manual analyst work each automated stage replaces (env-overridable).
MINUTES_PER_STAGE = {
    "dd": MINUTES_PER_CHECKLIST,
    "underwriting": int(os.environ.get("MINUTES_PER_UNDERWRITING", "120")),
    "comps": int(os.environ.get("MINUTES_PER_COMPS", "90")),
    "summary": int(os.environ.get("MINUTES_PER_SUMMARY", "45")),
}
STAGE_ORDER = ["dd", "underwriting", "comps", "summary"]
STAGE_LABELS = {"dd": "Due diligence", "underwriting": "Underwriting",
                "comps": "Rent comps", "summary": "One-pager summary"}
_counter_lock = threading.Lock()
_index_lock = threading.Lock()
_device_lock = threading.Lock()

# Schema metadata for labelling / grouping results in the UI.
_schema = yaml.safe_load((ROOT / "canonical" / "schema.yaml").read_text())
FIELD_BY_ID = {f["id"]: f for f in _schema["fields"]}
SECTIONS = [{"id": s["id"], "label": s["label"]} for s in _schema["sections"]]

# --------------------------------------------------------------------------- #
# Source catalog — what each automated answer is sourced from, by level.
# Reference content for the "Sources" tab. Three levels: federal/national (any
# US/CA location), California statewide (any CA parcel), and local/jurisdictional
# (varies by city & county — active: City of LA, City of San Diego + county
# parcels). Mirrors the reader registries in build/collect.py + the San Diego
# expansion (see SAN_DIEGO_EXPANSION.md).
# --------------------------------------------------------------------------- #
SOURCE_CATALOG = {
    "tiers": [
        {
            "key": "federal", "label": "Federal / national",
            "blurb": "Keyed off census tract, ZIP, or lat/long — identical answer anywhere in the U.S. No jurisdiction routing.",
            "groups": [
                {"source": "U.S. Census Bureau", "detail": "Geocoder + Incorporated Places",
                 "fields": ["Address (matched)", "County", "City / jurisdiction"],
                 "tracker": "https://geocoding.geo.census.gov/geocoder/"},
                {"source": "HUD", "detail": "QCT/DDA + Public Housing Authority layers",
                 "fields": ["Qualified Census Tract (QCT)", "Difficult Development Area (DDA)", "Public Housing Authority (PHA)"],
                 "tracker": "https://www.huduser.gov/portal/sadda/sadda_qct.html"},
                {"source": "FEMA", "detail": "National Flood Hazard Layer (NFHL)",
                 "fields": ["Flood zone"],
                 "tracker": "https://msc.fema.gov/portal/home"},
                {"source": "U.S. Treasury / CDFI Fund", "detail": "Opportunity Zone designations",
                 "fields": ["Opportunity Zone"],
                 "tracker": "https://opportunityzones.hud.gov/resources/map"},
                {"source": "USGS", "detail": "3DEP / EPQS elevation",
                 "fields": ["Slope grade"],
                 "tracker": "https://apps.nationalmap.gov/viewer/"},
                {"source": "FCC / HIFLD", "detail": "Cellular tower registry",
                 "fields": ["Cell towers"],
                 "tracker": "https://wireless2.fcc.gov/UlsApp/AsrSearch/asrRegistrationSearch.jsp"},
                {"source": "OpenStreetMap", "detail": "Overpass — open community data",
                 "fields": ["Nearest bus stop, grocery, park, clinic, library, pharmacy, school & qualifying transit"],
                 "tracker": "https://www.openstreetmap.org/"},
            ],
        },
        {
            "key": "state", "label": "California — statewide",
            "blurb": "State agency datasets covering every CA parcel — these port across counties unchanged.",
            "groups": [
                {"source": "CTCAC / HCD", "detail": "Opportunity & AFFH maps (statewide, all 58 counties)",
                 "fields": ["Resource area", "Neighborhood change area"],
                 "tracker": "https://www.treasurer.ca.gov/ctcac/opportunity.asp"},
                {"source": "CDLAC", "detail": "Geographic region lookup",
                 "fields": ["Geographic pool (region)"],
                 "tracker": "https://www.treasurer.ca.gov/cdlac/"},
                {"source": "CAL FIRE / OSFM", "detail": "Fire Hazard Severity Zones",
                 "fields": ["Very high fire hazard zone"],
                 "tracker": "https://egis.fire.ca.gov/FHSZ/"},
                {"source": "CA Coastal Commission / Caltrans", "detail": "Coastal Zone boundary",
                 "fields": ["Coastal zone"],
                 "tracker": "https://www.coastal.ca.gov/maps/"},
                {"source": "CalGEM", "detail": "WellSTAR oil & gas wells",
                 "fields": ["Wells on site"],
                 "tracker": "https://maps.conservation.ca.gov/calgem/findwells/"},
                {"source": "California Geological Survey", "detail": "EQ Zapp seismic hazard zones",
                 "fields": ["Liquefaction zone", "Alquist-Priolo fault zone"],
                 "tracker": "https://www.conservation.ca.gov/cgs/geohazards/eq-zapp"},
                {"source": "SWRCB GeoTracker", "detail": "UST/LUST cleanup sites (EPA fallback)",
                 "fields": ["Underground storage tanks"],
                 "tracker": "https://geotracker.waterboards.ca.gov/map/"},
            ],
        },
    ],
    "local": {
        "label": "Local / jurisdictional",
        "blurb": ("Zoning, parcel and entitlement data come from each jurisdiction's own GIS, so the source "
                  "varies by city & county. Active: City of Los Angeles (ZIMAS), unincorporated LA County "
                  "(County DRP), and City of San Diego (plus county parcel layers). Unincorporated parcels "
                  "are governed by County Title 22, not the City LAMC. Parcels in other incorporated cities "
                  "route to manual review."),
        "cols": ["Field", "Los Angeles (City)", "Unincorporated LA County", "San Diego"],
        "trackers": [
            {"label": "City of LA — ZIMAS", "url": "https://zimas.lacity.org/"},
            {"label": "Unincorporated LA County — GIS-NET Public", "url": "https://planning.lacounty.gov/maps-and-gis/gis-net-public/"},
            {"label": "San Diego — SanGIS parcel map", "url": "https://www.sangis.org/pages/interactive-map"},
            {"label": "San Diego — City zoning grid", "url": "https://www.sandiego.gov/development-services/zoning-maps/grid-map"},
        ],
        "rows": [
            {"field": "APN / parcel ID", "la": "LA City / County Parcels", "county": "LA County Assessor parcels", "sd": "SANDAG County Parcels"},
            {"field": "Land area (SF)", "la": "LA City / County Parcels (EPSG:2229)", "county": "LA County Assessor (EPSG:2229)", "sd": "SANDAG County Parcels (EPSG:2230)"},
            {"field": "Zoning", "la": "ZIMAS / NavigateLA", "county": "County DRP — Zoning (Title 22) + General Plan", "sd": "City of San Diego — Base Zones"},
            {"field": "Specific plan / overlay", "la": "ZIMAS / NavigateLA", "county": "County DRP — SP zone / CSD / Zoned District / SEA", "sd": "City of San Diego — DSD Zoning Overlay"},
            {"field": "Council / supervisor district", "la": "ZIMAS / NavigateLA", "county": "County — Supervisorial District", "sd": "City of San Diego — DoIT public layers"},
            {"field": "Historic status", "la": "ZIMAS / SurveyLA", "county": "Manual — no County REST layer", "sd": "City of San Diego — Historic Preservation"},
            {"field": "TOC / transit tier", "la": "LA City Planning — TOC", "county": "County DRP — Transit Oriented District (TOD)", "sd": "SD Transit Priority Area"},
            {"field": "½-mile major transit", "la": "LA City Planning", "county": "Derived from County TOD (JUDGMENT)", "sd": "SD Transit Priority Area (SB 743)"},
            {"field": "Airport hazard zone", "la": "LA County ALUC (A-NET)", "county": "LA County ALUC (A-NET)", "sd": "City of San Diego — DSD Airports (ALUC)"},
            {"field": "Q conditions", "la": "ZIMAS / NavigateLA", "county": "N/A — LA-City-only zoning concept", "sd": "N/A — LA-only zoning concept"},
            {"field": "Methane hazard zone", "la": "ZIMAS / NavigateLA", "county": "N/A — LA-City-only zoning concept", "sd": "N/A — LA-only zoning concept"},
            {"field": "Transitional height", "la": "Derived (LAMC)", "county": "N/A — LA-City-only (LAMC) concept", "sd": "N/A — LA-only zoning concept"},
        ],
    },
}

MAX_JOBS = 50                          # keep the most recent N; prune older + their files
_jobs = {}
_lock = threading.Lock()

# Reader landing states that are legitimate "good" outcomes (mirrors runner.py).
_LANDING = {"VERIFIED", "JUDGMENT", "NA", "COMPUTED", "OM-SOURCED"}


def _now():
    # timezone-aware UTC so timestamps are unambiguous; the UI renders them in
    # Pacific on a 12-hour clock (see app.js fmtWhen).
    return datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds")


def _jsonable(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _field_meta(fid):
    f = FIELD_BY_ID.get(fid, {})
    return f.get("label", fid), f.get("section", "other")


def _record(job, fid, answer, state, notes):
    label, section = _field_meta(fid)
    entry = {
        "id": fid, "label": label, "section": section,
        "answer": _jsonable(answer), "state": state, "notes": notes or "",
    }
    # The worker thread writes while poll requests read; guard the dict so a
    # concurrent iteration in public_view() can't hit "changed size".
    with job["_lock"]:
        job["fields"][fid] = entry


# --------------------------------------------------------------------------- #
# OM reconciliation — default to the OM; switch to a DD answer only when the DD
# process produced a different, cited value (then flag the discrepancy).
# --------------------------------------------------------------------------- #
import re as _re


def _digits(s):
    return _re.sub(r"\D", "", str(s or ""))


def _num(s):
    m = _re.search(r"-?\d[\d,]*(?:\.\d+)?", str(s or ""))
    return float(m.group(0).replace(",", "")) if m else None


def _values_agree(fid, om_val, dd_val):
    """Loose agreement test between an OM value and a DD reader value, by field type."""
    if dd_val is None:
        return False
    if fid == "land_sf":
        a, b = _num(om_val), _num(dd_val)
        if a is None or b is None:
            return False
        return abs(a - b) <= max(300.0, 0.02 * max(a, b))      # within 2% or 300 sf
    if fid == "apn":
        return _digits(om_val) == _digits(dd_val) and bool(_digits(om_val))
    if fid == "address":
        on, dn = _digits(om_val.split()[0] if om_val else ""), _digits(dd_val.split()[0] if dd_val else "")
        return bool(on) and on == dn                            # same house number = good enough
    if fid in ("county", "city_jurisdiction"):
        a, b = str(om_val).lower(), str(dd_val).lower()
        return a in b or b in a
    return str(om_val).strip().lower() == str(dd_val).strip().lower()


def _mutate_field(job, fid, **changes):
    with job["_lock"]:
        if fid in job["fields"]:
            job["fields"][fid].update(changes)


def _apply_om_merge(job, om_fields, ws, outcomes):
    """Reconcile extracted OM values against DD outcomes and write them to the workbook.

    Rule: default to the OM. When a DD reader produced a value too, the DD answer
    wins on a real conflict (it has a cited source) and the cell is flagged
    JUDGMENT; on agreement the DD VERIFIED value stands with an 'OM agrees' note.
    Where DD has no value (no reader, or it failed), the OM value fills the cell
    as OM-SOURCED."""
    def put(row, col, val):
        ws.cell(row, col, val)

    def append_note(row, text):
        prev = ws.cell(row, 5).value
        ws.cell(row, 5, (prev + " | " if prev else "") + text)
        return ws.cell(row, 5).value

    merged = []
    for f in om_fields:
        fid = f["field_id"]
        field = FIELD_BY_ID.get(fid)
        if not field:
            continue
        row = int(field["answer_cell"][1:])
        omv, q, conf = f["value"], f["source_quote"], f["confidence"]
        dd = outcomes.get(fid)
        dd_ok = bool(dd) and dd[0] == "ok"
        dd_val = dd[1].get("answer") if dd_ok else None

        if dd_ok and _values_agree(fid, omv, dd_val):
            note = f'OM agrees ({conf}): "{omv}". OM source: "{q}".'
            full = append_note(row, note)
            _mutate_field(job, fid, notes=full, om="agree")
            merged.append({**f, "outcome": "agree", "dd_value": _jsonable(dd_val)})
        elif dd_ok:
            note = (f'CONFLICT — OM stated "{omv}" but DD (cited) found "{dd_val}"; DD value used. '
                    f'OM source: "{q}".')
            ws.cell(row, 1, "JUDGMENT")
            full = append_note(row, note)
            _mutate_field(job, fid, state="JUDGMENT", notes=full, om="conflict")
            merged.append({**f, "outcome": "conflict", "dd_value": _jsonable(dd_val)})
        else:
            reason = "no DD source for this field" if dd is None else "DD reader failed"
            note = f'From OM ({conf} confidence) — {reason}, OM value used. Source: "{q}".'
            put(row, 1, "OM-SOURCED")
            put(row, 3, omv)
            ws.cell(row, 5, note)
            label, section = _field_meta(fid)
            with job["_lock"]:
                job["fields"][fid] = {"id": fid, "label": label, "section": section,
                                      "answer": _jsonable(omv), "state": "OM-SOURCED",
                                      "notes": note, "om": "sourced"}
            merged.append({**f, "outcome": "om-sourced", "dd_value": None})
    return merged


# --------------------------------------------------------------------------- #
# single address
# --------------------------------------------------------------------------- #
def _display_outcome(outcome):
    """Map a run_reader result to (state, answer, notes) for the UI.

    A fresh single run = one attempt, so a failure shows as TOOL-FAIL (the saved
    workbook records the real 'TOOL-FAIL 1/3' counter via apply_outcome)."""
    kind, payload = outcome
    if kind == "ok":
        st = payload.get("state", "VERIFIED")
        if st not in _LANDING:
            st = "VERIFIED"
        return st, payload.get("answer"), payload.get("notes", "")
    return "TOOL-FAIL", None, str(payload)[:200]


def run_single(job):
    # OM (optional) — extract deal facts first; an uploaded OM can also supply the address.
    om_fields = []
    if job["input"].get("om_bytes"):
        job["phase"] = "Reading the Offering Memorandum with Gemini…"
        try:
            import om_extract
            om_fields = om_extract.extract(job["input"]["om_bytes"],
                                           job["input"].get("om_name") or "om.pdf")
            job["om"] = {"name": job["input"].get("om_name"), "extracted": om_fields, "error": None}
        except Exception as e:
            job["om"] = {"name": job["input"].get("om_name"), "extracted": [], "error": str(e)}

    address = (job["input"].get("address") or "").strip()
    if not address:
        address = next((f["value"] for f in om_fields if f["field_id"] == "address"), "")
    if not address:
        raise RuntimeError("No address provided, and none could be read from the OM.")

    # A ';'-separated address runs as one site (assemblage): point/tract readers
    # use the primary (first) parcel; land_sf is summed + APNs listed across all.
    addresses = _collect._parse_addresses(address)
    addresses = list(addresses) if isinstance(addresses, (list, tuple)) else [addresses]
    geos = [geocode(a) for a in addresses]
    geo = geos[0]                          # primary parcel drives point/tract readers
    multi = len(geos) > 1
    job["geo"] = {
        "matched_address": geo["matched_address"], "geoid": geo["geoid"],
        "lat": round(geo["lat"], 6), "lon": round(geo["lon"], 6),
    }
    job["label"] = (f"{geo['matched_address']} (+{len(geos) - 1} parcels)"
                    if multi else geo["matched_address"])

    active = dict(_collect.READERS)
    if multi:                              # parcel fields are aggregated across the assemblage, not snapped to primary
        for fid in ("address", "apn", "land_sf"):
            active.pop(fid, None)
    in_la = False
    block = None        # which municipal block ran: "la_city" | "la_county" | "san_diego" | None
    try:
        in_la = zimas.in_la_city(geo)      # also warms the shared parcel snap
    except Exception:
        pass
    if in_la:
        active.update(_collect.ZIMAS_READERS)
        block = "la_city"
    else:
        try:
            if _county_basename(geo) == "Los Angeles" and lacounty.is_unincorporated(geo):
                active.update(_collect.LACOUNTY_READERS)
                block = "la_county"
            elif _county_basename(geo) == "San Diego":
                active.update(_collect.SD_READERS)
                block = "san_diego"
        except Exception:
            pass
    job["in_la_city"] = in_la
    job["block"] = block
    job["total"] = len(active) + (3 if multi else 0)   # +3 aggregated parcel fields
    _block_phase = {
        "la_city": "In City of LA — running ZIMAS zoning/hazard block. ",
        "la_county": "In unincorporated LA County — running LA County (DRP) zoning/land-use block. ",
        "san_diego": "In the City of San Diego — running SD municipal block. ",
    }.get(block, "Outside LA City / unincorporated LA County / San Diego — municipal block skipped. ")
    job["phase"] = (f"Assemblage of {len(geos)} addresses — " if multi else "") + \
                   _block_phase + "Running readers…"

    try:
        nc._load()                         # warm the Neighborhood-Change cache once
    except Exception:
        pass

    # Phase 1 — fan all readers out; report each as it finishes.
    outcomes = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = {ex.submit(run_reader, (lambda fn=fn: fn(geo))): fid
                for fid, fn in active.items()}
        for fut in as_completed(futs):
            fid = futs[fut]
            try:
                outcome = fut.result()
            except Exception as e:               # defensive — run_reader catches its own
                outcome = ("fail", e)
            outcomes[fid] = outcome
            st, ans, notes = _display_outcome(outcome)
            _record(job, fid, ans, st, notes)
            job["completed"] = len(outcomes)

    # Assemblage — aggregate the parcel fields across all addresses + flag tract divergence.
    if multi:
        job["phase"] = f"Combining {len(geos)} parcels (summing land area, listing APNs)…"
        agg, parts = _collect._assemble_parcels(geos)
        tracts = sorted({g["geoid"] for g in geos})
        if len(tracts) > 1 and agg["address"][0] == "ok":
            agg["address"][1]["notes"] += (
                f" ASSEMBLAGE SPANS {len(tracts)} CENSUS TRACTS ({', '.join(tracts)}) — "
                f"tract-based fields (QCT/DDA/resource/opportunity zone/neighborhood change) "
                f"reflect the PRIMARY parcel only; verify per parcel.")
        outcomes.update(agg)
        for fid in agg:
            st, ans, notes = _display_outcome(agg[fid])
            _record(job, fid, ans, st, notes)
        job["parcels"] = [{"apn": p["apn"] or "(unresolved)", "n_lots": 1,
                           "land_sf": p["area"], "geoid": p["geoid"]} for p in parts]
        job["combined_sf"] = agg["land_sf"][1]["answer"] if agg["land_sf"][0] == "ok" else None
        job["completed"] = len(outcomes)

    # Phase 2 — write the workbook once (reuses apply_outcome for the real file).
    wb = load_workbook(TEMPLATE)
    ws, log = wb["Site DD"], wb["State Log"]
    ts = _now()
    for fid in outcomes:
        apply_outcome(ws, log, FIELD_BY_ID[fid], outcomes[fid],
                      property_id=job["input"].get("property_id") or "WEB", ts=ts)

    # Reconcile OM deal facts over the DD results (default OM; DD wins cited conflicts).
    if om_fields:
        job["phase"] = "Reconciling OM values with DD findings…"
        merged = _apply_om_merge(job, om_fields, ws, outcomes)
        if job.get("om"):
            job["om"]["merged"] = merged

    out_path = RUN_DIR / f"{job['id']}.xlsx"
    wb.save(out_path)
    job["file"] = str(out_path)
    job["filename"] = _safe_name(geo["matched_address"]) + ".xlsx"
    job["phase"] = "Complete"


# --------------------------------------------------------------------------- #
# multi-APN assemblage
# --------------------------------------------------------------------------- #
def run_assemblage(job):
    apns = job["input"]["apns"]
    job["phase"] = f"Resolving {len(apns)} APN(s) and running readers across parcels…"
    out_path = RUN_DIR / f"{job['id']}.xlsx"
    shutil.copy(TEMPLATE, out_path)

    result = _assemblage.assess(str(out_path), apns, property_id="WEB-ASSEMBLAGE")

    job["parcels"] = [
        {"apn": p["apn"], "n_lots": p["n_lots"], "land_sf": p["land_sf"], "geoid": p["geoid"]}
        for p in result["parcels"]
    ]
    job["combined_sf"] = result["combined_sf"]

    # Site-level fields assess() computes itself (written to the workbook, not in .fields).
    apn_list = ", ".join(p["apn"] for p in result["parcels"])
    _record(job, "apn", apn_list, "VERIFIED",
            f"Block assemblage of {len(result['parcels'])} APN(s).")
    _record(job, "land_sf", round(result["combined_sf"]), "VERIFIED",
            f"Combined gross land area = {result['combined_sf']:,.1f} sf "
            f"({result['combined_sf'] / 43560:.3f} ac) across {len(result['parcels'])} APN(s).")
    for fid, a in result["fields"].items():
        _record(job, fid, a["answer"], a["state"], a.get("notes", ""))

    job["total"] = job["completed"] = len(job["fields"])
    job["file"] = str(out_path)
    job["filename"] = f"assemblage_{result['parcels'][0]['apn'].replace('-', '')}_{len(apns)}APN.xlsx"
    job["phase"] = "Complete"


# --------------------------------------------------------------------------- #
# financial model — DD checklist -> Stick + Modular pro-forma (.xlsm, zipped)
# --------------------------------------------------------------------------- #
def run_underwrite(job):
    inp = job["input"]

    # Source DD workbook: an uploaded .xlsx, or a prior DD run's saved workbook.
    if inp.get("dd_bytes"):
        dd_path = RUN_DIR / f"{job['id']}_dd.xlsx"
        dd_path.write_bytes(inp["dd_bytes"])
    elif inp.get("from_job"):
        prior = get_job(inp["from_job"])
        if not prior or not prior.get("file"):
            raise RuntimeError("Source checklist not found (it may have expired) — re-run the DD or upload the file.")
        dd_path = Path(prior["file"])
        job["label"] = (prior.get("label") or "") + " — financial model"
    else:
        raise RuntimeError("No DD checklist provided.")

    job["phase"] = "Reading the DD checklist…"
    dd = _underwrite.read_dd(dd_path)

    # Review/edit step: fold any analyst overrides into the DD before projecting.
    import uw_logic as _uwl
    deal_override = None
    if inp.get("overrides"):
        dd, deal_override = _uwl.apply_overrides(dd, inp["overrides"])

    # Build the v28 LIHTC financial model (Modular + Stick construction) from the
    # DD facts. deal name + meta come from base_cells (what export() used too), so
    # we no longer generate the legacy Stick + Modular template models at all.
    job["phase"] = "Building the v28 LIHTC financial model…"
    _base, meta = _uwl.base_cells(dd)
    deal = deal_override or inp.get("name") or _base.get(("Pro_Forma", "B2")) or "Untitled Deal"
    job["label"] = f"{deal} — financial model"

    # Deal-type selector: LIHTC (default, unchanged) vs non-LIHTC (market/mixed).
    if (inp.get("deal_type") or "lihtc").lower() in ("nonlihtc", "non-lihtc", "market"):
        return _run_underwrite_nonlihtc(job, dd, deal, meta)

    import modularz_calc as _mz
    v28_files = []  # (arcname, bytes)
    for _m in ("Modular", "Stick"):
        v28_files.append((f"{deal} — LIHTC v28 ({_m}).xlsm",
                          _mz.build_from_dd(dd, method=_m, deal_name=deal)))

    import zipfile
    zip_path = RUN_DIR / f"{job['id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for arc, blob in v28_files:
            z.writestr(arc, blob)

    job["underwrite"] = {
        "deal": deal,
        "product": meta.get("product"),
        "resource": meta.get("resource_mapped"),
        "flags": meta.get("flags", []),
        "default_model": v28_files[0][0],
        "models": [arc for arc, _ in v28_files],
        "inputs": {k: _jsonable(v) for k, v in dd.items()},
        "envelope_defaults": {
            "residential_stories": _jsonable(dd.get("residential_stories")) or _uwl.DEFAULT_STORIES,
            "building_nrsf": _jsonable(dd.get("building_nrsf")) or _uwl.DEFAULT_NRSF,
        },
        "hand_fields": ["BIPOC", "Prevailing wage"],
    }
    job["total"] = job["completed"] = 1
    job["file"] = str(zip_path)
    job["filename"] = f"{_safe_name(deal)}_models.zip"
    job["phase"] = "Complete"


def _run_underwrite_nonlihtc(job, dd, deal, meta):
    """Non-LIHTC (market/mixed) financial model from DD facts + a unit program +
    comp rents (+ optional T-12 OpEx). Runs the clean pre-v28 ModularZ market
    engine via nonlihtc_calc; the LIHTC v30 path is untouched. The unit program
    and rents come from the review step / comp scraper via inp['nonlihtc']."""
    inp = job["input"]
    params = inp.get("nonlihtc") or {}
    units_by_bed = params.get("units_by_bed")
    rents_by_bed = params.get("rents_by_bed")
    if not units_by_bed or not rents_by_bed:
        raise RuntimeError(
            "Non-LIHTC model needs a unit program (units_by_bed) and market rents "
            "(rents_by_bed) from the review step / comp scraper.")

    job["phase"] = "Building the non-LIHTC market model…"
    import nonlihtc_calc as _nl
    land_cost = params.get("land_cost") or _nl._parse_num(dd.get("acquisition_price"))
    blob = _nl.build_from_dd(
        dd, units_by_bed=units_by_bed, rents_by_bed=rents_by_bed,
        opex=params.get("opex"), financing=params.get("financing"),
        land_cost=land_cost, deal_name=deal,
    )

    # Headline returns for the UI (recalc the same friendly inputs). Best-effort:
    # the .xlsx is the deliverable, so a recalc hiccup must not fail the job.
    returns = {}
    try:
        friendly = _nl.build_market_inputs(
            units_by_bed=units_by_bed, rents_by_bed=rents_by_bed,
            land_cost=land_cost or 0, opex=params.get("opex"),
            financing=params.get("financing"),
        )
        returns = _nl.recalc(friendly)
    except Exception:  # noqa: BLE001
        traceback.print_exc()

    import zipfile
    arc = f"{deal} — Non-LIHTC (Market).xlsx"
    zip_path = RUN_DIR / f"{job['id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(arc, blob)

    job["underwrite"] = {
        "deal": deal,
        "deal_type": "nonlihtc",
        "models": [arc],
        "default_model": arc,
        "returns": {k: _jsonable(v) for k, v in returns.items()},
        "inputs": {k: _jsonable(v) for k, v in dd.items()},
        "hand_fields": [],
    }
    job["total"] = job["completed"] = 1
    job["file"] = str(zip_path)
    job["filename"] = f"{_safe_name(deal)}_nonlihtc.zip"
    job["phase"] = "Complete"


# --------------------------------------------------------------------------- #
# LIHTC scenario batch — selected scenario list -> one .xlsm each -> zip
# --------------------------------------------------------------------------- #
def run_lihtc_scenarios(job):
    inp = job["input"]

    if inp.get("dd_bytes"):
        dd_path = RUN_DIR / f"{job['id']}_dd.xlsx"
        dd_path.write_bytes(inp["dd_bytes"])
    elif inp.get("from_job"):
        prior = get_job(inp["from_job"])
        if not prior or not prior.get("file"):
            raise RuntimeError("Source checklist not found (it may have expired) — re-run the DD or upload the file.")
        dd_path = Path(prior["file"])
        job["label"] = (prior.get("label") or "") + " — LIHTC scenarios"
    else:
        raise RuntimeError("No DD checklist provided.")

    job["phase"] = "Reading the DD checklist…"
    dd = _underwrite.read_dd(dd_path)

    import uw_logic as _uwl
    deal_override = None
    if inp.get("overrides"):
        dd, deal_override = _uwl.apply_overrides(dd, inp["overrides"])

    _base, _meta = _uwl.base_cells(dd)
    deal = deal_override or inp.get("name") or _base.get(("Pro_Forma", "B2")) or "Untitled Deal"
    job["label"] = f"{deal} — LIHTC scenarios"

    selected = inp.get("scenarios") or []
    if not selected:
        raise RuntimeError("No scenarios were selected.")

    job["total"] = len(selected)
    job["completed"] = 0

    import modularz_calc as _mz

    def _build(scn):
        return (f"{deal} — {scn['name']}.xlsm", _mz.build_for_scenario(dd, scn, deal_name=deal))

    job["phase"] = "Building scenario models…"
    with ThreadPoolExecutor(max_workers=len(selected)) as ex:
        build_futs = [ex.submit(_build, scn) for scn in selected]
        scn_files = []
        for fut in as_completed(build_futs):
            scn_files.append(fut.result())
            job["completed"] += 1
            job["phase"] = f"Built {job['completed']}/{len(selected)} models…"

    import zipfile
    zip_path = RUN_DIR / f"{job['id']}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for arc, blob in scn_files:
            z.writestr(arc, blob)

    job["file"] = str(zip_path)
    job["filename"] = f"{_safe_name(deal)}_scenarios.zip"
    job["total"] = job["completed"] = len(scn_files)
    job["phase"] = "Complete"


# --------------------------------------------------------------------------- #
# pdf_summary — generates the one-pager PDF from a completed lihtc_scenarios job
# --------------------------------------------------------------------------- #
def run_pdf_summary(job):
    inp = job["input"]
    from_jid = inp.get("from_job")
    if not from_jid:
        raise RuntimeError("No source scenario job provided.")

    prior = get_job(from_jid)
    if not prior:
        raise RuntimeError("Source scenario job not found.")

    # Recover the DD path and inputs from the prior scenario job.
    prior_inp = prior.get("input") or {}
    if prior_inp.get("dd_bytes"):
        dd_path = RUN_DIR / f"{from_jid}_dd.xlsx"
        if not dd_path.exists():
            dd_path.write_bytes(prior_inp["dd_bytes"])
    elif prior_inp.get("from_job"):
        grandparent = get_job(prior_inp["from_job"])
        if not grandparent or not grandparent.get("file"):
            raise RuntimeError("Original DD checklist not found — re-run the DD.")
        dd_path = Path(grandparent["file"])
    else:
        raise RuntimeError("Cannot locate the original DD checklist.")

    job["phase"] = "Reading the DD checklist…"
    dd = _underwrite.read_dd(dd_path)

    import uw_logic as _uwl
    overrides = prior_inp.get("overrides") or {}
    deal_override = None
    if overrides:
        dd, deal_override = _uwl.apply_overrides(dd, overrides)

    _base, _meta = _uwl.base_cells(dd)
    deal = deal_override or prior_inp.get("name") or _base.get(("Pro_Forma", "B2")) or "Untitled Deal"
    job["label"] = f"{deal} — PDF summary"

    selected = prior_inp.get("scenarios") or []
    if not selected:
        raise RuntimeError("No scenarios found on the source job.")

    import modularz_calc as _mz

    job["total"] = len(selected)
    job["completed"] = 0
    job["phase"] = f"Computing outputs for {len(selected)} scenarios…"

    def _calc(scn):
        try:
            num, txt = _mz._scenario_inputs(dd, scn, deal_name=deal)
            vals = _mz.recalc_isolated(num, txt, read=_mz.SCENARIO_READ)
        except Exception:
            vals = {}
        tdc   = vals.get("D55")
        units = vals.get("C14")
        return {
            "name":    scn["name"],
            "constr":  scn["constr"],
            "stories": scn["stories"],
            "lf":      scn.get("lf", "No"),
            "results": {
                "units":     int(round(units)) if units else None,
                "nrsf":      vals.get("C17"),
                "tdc":       tdc,
                "tdc_unit":  round(tdc / units) if (tdc and units) else None,
                "irr":       vals.get("C25"),
                "moic":      vals.get("C26"),
                "perm_loan": vals.get("D75"),
                "tc_equity": vals.get("D80"),
            },
        }

    with ThreadPoolExecutor(max_workers=min(len(selected), 4)) as ex:
        calc_futs = [ex.submit(_calc, scn) for scn in selected]
        scn_data  = []
        for fut in as_completed(calc_futs):
            scn_data.append(fut.result())
            job["completed"] += 1
            job["phase"] = f"Computed {job['completed']}/{len(selected)} scenarios…"

    # Preserve the original selection order for the PDF table.
    order = {scn["name"]: i for i, scn in enumerate(selected)}
    scn_data.sort(key=lambda s: order.get(s["name"], 999))

    job["phase"] = "Generating PDF…"
    import sys, importlib
    sys.path.insert(0, str(Path(__file__).parent.parent / "build" / "sources"))
    try:
        import deal_onepager as _op
        importlib.reload(_op)
    except ImportError:
        import deal_onepager as _op

    pdf_deal = _op.deal_from_scenarios(
        name=deal,
        address=dd.get("address") or "",
        dd=dd,
        overrides=overrides,
        scenarios_with_data=scn_data,
        as_of=datetime.datetime.now(datetime.timezone.utc).date(),
    )
    pdf_path = RUN_DIR / f"{job['id']}_onepager.pdf"
    _op.generate(pdf_deal, str(pdf_path))

    job["file"]     = str(pdf_path)
    job["filename"] = f"{_safe_name(deal)}_one_pager.pdf"
    job["total"]    = job["completed"] = len(selected)
    job["phase"]    = "Complete"


# --------------------------------------------------------------------------- #
# rent comps — shortlist (run) -> review/edit matrix -> formatted CTCAC grid
# --------------------------------------------------------------------------- #
def run_comps(job):
    import comps as _comps
    inp = job["input"]
    addr = (inp.get("address") or "").strip()
    beds = inp.get("beds") or [0, 1, 2]
    # Source priority: ai (Firecrawl+Gemini) > rentcast > demo
    has_ai = bool(os.environ.get("FIRECRAWL_API_KEY") and os.environ.get("GEMINI_API_KEY"))
    has_rentcast = bool(os.environ.get("RENTCAST_API_KEY"))
    if has_ai:
        source, demo = "ai", False
        job["phase"] = f"Scraping Zillow for beds {beds} via Firecrawl + Gemini…"
    elif has_rentcast:
        source, demo = "rentcast", False
        job["phase"] = f"Geocoding subject and shortlisting comps for beds {beds}…"
    else:
        source, demo = "ai", True
        job["phase"] = f"Demo data (no API keys set) — beds {beds}…"
    geo, by_bed = _comps.collect_comps(addr, beds, _comps.rentcast.DEFAULT_RADIUS_MI,
                                       inp.get("top", 4), demo=demo, use_avm=(source == "rentcast" and not demo),
                                       source=source)
    job["geo"] = {"matched_address": geo["matched_address"],
                  "lat": round(geo["lat"], 6), "lon": round(geo["lon"], 6)}
    job["label"] = geo["matched_address"] + " — rent comps"
    # keep the rollup records for the editor + grid write
    job["comps_data"] = {b: payload["comps"] for b, payload in by_bed.items()}
    n = sum(len(v) for v in job["comps_data"].values())
    job["total"] = job["completed"] = n
    job["phase"] = f"Shortlisted {n} comp(s) across {len(beds)} bed type(s). Open the editor to adjust."


def comps_intake(jid):
    """Subject + comp rows + the adjustment ruleset for the comp review/edit matrix."""
    job = get_job(jid)
    if not job or job.get("kind") != "comps" or not job.get("comps_data"):
        raise ValueError("Comp run not found (it may have expired) — re-run the comps.")
    import comp_adjust as CA
    ruleset = {
        "age_per_year": CA.AGE_PER_YEAR, "size_rate_fraction": CA.SIZE_RATE_FRACTION,
        "guardrail": CA.GUARDRAIL,
        "amenity_labels": CA.AMENITY_LABELS,
        "amenity_values": {l: CA.amenity_value(l) for l in CA.AMENITY_LABELS},
        "utility_labels": CA.UTILITY_LABELS, "utility_values": CA.UTILITY_VALUES,
    }
    beds = []
    for b, rows in sorted(job["comps_data"].items()):
        beds.append({"bed": b,
                     "comps": [{"address": c.get("address"), "city": c.get("city"),
                                "distance_mi": c.get("distance_mi"),
                                "sf": c.get("unit_size_sf"), "rent": c.get("base_rent"),
                                "year": c.get("year_built"), "baths": c.get("bathrooms"),
                                "amenity_flags": c.get("amenity_flags") or {},
                                "utility_flags": c.get("utility_flags") or {}}
                               for c in rows]})
    return {"label": job.get("label") or jid, "geo": job.get("geo"),
            "beds": beds, "ruleset": ruleset}


def run_comps_grid(job):
    """Write the formatted CTCAC grid from the editor's edited subject + comp chars."""
    import comps as _comps
    inp = job["input"]
    prior = get_job(inp.get("from_job") or "")
    geo = (prior.get("geo") if prior else None) or {"matched_address": inp.get("address") or "Subject"}
    grid = inp.get("grid") or {}     # {bed: {subject:{...}, comps:[{...}]}}
    by_bed, subjects = {}, {}
    for bed_str, data in grid.items():
        b = int(bed_str)
        comps_list = [{"address": c.get("address"), "city": c.get("city"),
                       "distance_mi": c.get("distance_mi"), "unit_size_sf": c.get("sf"),
                       "base_rent": c.get("rent"), "value_ratio": (round(c["rent"] / c["sf"], 2)
                       if c.get("rent") and c.get("sf") else None), "bedrooms": b,
                       "bathrooms": c.get("baths"), "year_built": c.get("year"),
                       "_amenities": c.get("amenities", {}), "_utilities": c.get("utilities", {})}
                      for c in data.get("comps", [])]
        by_bed[b] = {"comps": comps_list, "estimate": None}
        s = data.get("subject", {})
        subjects[b] = {"sf": s.get("sf"), "rent": s.get("rent"), "year": s.get("year"),
                       "baths": s.get("baths"), "city": s.get("city"), "m_or_l": "M",
                       "amenities": s.get("amenities", {}), "utilities": s.get("utilities", {})}
    # carry comp amenity/utility chars into the engine via the writer's ecomps map
    _orig = _comps.write_ctcac_grid
    out_path = RUN_DIR / f"{job['id']}.xlsx"
    job["phase"] = "Writing the formatted CTCAC rent-comp grid…"
    _comps.write_ctcac_grid(geo, by_bed, str(out_path), subjects,
                            comp_chars={b: [{"amenities": c["_amenities"], "utilities": c["_utilities"]}
                                            for c in by_bed[b]["comps"]] for b in by_bed})
    job["file"] = str(out_path)
    job["filename"] = _safe_name(geo["matched_address"]) + " — rent comp grid.xlsx"
    job["total"] = job["completed"] = 1
    job["phase"] = "Complete"


# --------------------------------------------------------------------------- #
# review/edit step — editable model inputs from a completed DD run
# --------------------------------------------------------------------------- #
def _norm_addr(s):
    """Normalize an address for loose matching: lowercase, drop the assemblage
    suffix (e.g. '(+2 parcels)'), collapse non-alphanumerics to single spaces."""
    import re
    s = (s or "").lower()
    s = re.sub(r"\s*\(\+\d+\s*parcels?\)\s*$", "", s)
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def latest_comp_rents(address):
    """Median scraped market rent per bed from the most recent completed comps run
    whose subject matches `address`. Feeds the non-LIHTC review form's rent defaults
    straight from the AI comp scraper. Returns {rents_by_bed, counts, source_job,
    address} or None when no matching comp run exists this session."""
    import statistics
    key = _norm_addr(address)
    if not key:
        return None
    with _lock:
        runs = [j for j in _jobs.values()
                if j.get("kind") == "comps" and j.get("status") == "done"
                and j.get("comps_data")
                and _norm_addr((j.get("geo") or {}).get("matched_address")) == key]
    if not runs:
        return None
    runs.sort(key=lambda j: j.get("finished") or "", reverse=True)
    job = runs[0]
    rents, counts = {}, {}
    for b, rows in job["comps_data"].items():
        vals = [c.get("base_rent") for c in rows
                if isinstance(c.get("base_rent"), (int, float)) and c.get("base_rent")]
        if vals:
            rents[int(b)] = int(round(statistics.median(vals)))
            counts[int(b)] = len(vals)
    if not rents:
        return None
    return {"rents_by_bed": rents, "counts": counts, "source_job": job["id"],
            "address": (job.get("geo") or {}).get("matched_address")}


def underwrite_intake(jid):
    """The editable intake (defaults + options + derived preview) for a DD job's
    financial model. Raises ValueError if the source checklist is unavailable."""
    job = get_job(jid)
    if not job or not job.get("file") or not os.path.exists(job["file"]):
        raise ValueError("Source checklist not found (it may have expired) — re-run the DD.")
    if job["kind"] not in ("single", "assemblage"):
        raise ValueError("Only a DD checklist can seed a financial model.")
    import uw_logic as _uwl
    dd = _underwrite.read_dd(Path(job["file"]))
    payload = _uwl.intake(dd)
    payload["label"] = job.get("label") or jid
    # If the analyst already ran the comp scraper on this subject, surface the
    # median market rents so the non-LIHTC form can pre-fill rents_by_bed.
    cr = latest_comp_rents((job.get("geo") or {}).get("matched_address") or job.get("label"))
    if cr:
        payload["comp_rents"] = cr
    return payload


# --------------------------------------------------------------------------- #
# job store
# --------------------------------------------------------------------------- #
def _safe_name(s):
    keep = "".join(c if c.isalnum() or c in " -," else "_" for c in str(s))
    return ("_".join(keep.split()) or "site")[:80]


# --------------------------------------------------------------------------- #
# time-saved counter (persisted) + stats
# --------------------------------------------------------------------------- #
def _read_counts():
    """Per-kind completed-run counts, persisted to the volume. Migrates the legacy
    {"count": N} format (which only ever counted DD single/assemblage runs) by
    treating it as DD 'single' runs so the historical number carries forward."""
    try:
        data = json.loads(COUNTER_FILE.read_text())
    except Exception:
        return {}
    counts = data.get("counts")
    if isinstance(counts, dict):
        return {k: int(v) for k, v in counts.items()}
    legacy = int(data.get("count", 0))     # old single-integer format
    return {"single": legacy} if legacy else {}


def _bump_count(kind):
    """Record one completed run of `kind`. Also writes a legacy `count` (DD total)
    so older code/readers still see a sensible DD number."""
    with _counter_lock:
        counts = _read_counts()
        counts[kind] = counts.get(kind, 0) + 1
        dd = sum(counts.get(k, 0) for k, s in STAGE_OF_KIND.items() if s == "dd")
        try:
            COUNTER_FILE.write_text(json.dumps({"counts": counts, "count": dd}))
        except Exception:
            traceback.print_exc()
        return counts


def stats():
    counts = _read_counts()
    # group raw per-kind counts into pipeline stages
    by_stage = {s: 0 for s in MINUTES_PER_STAGE}
    for kind, n in counts.items():
        stage = STAGE_OF_KIND.get(kind)
        if stage:
            by_stage[stage] = by_stage.get(stage, 0) + int(n)
    by_stage["dd"] += STARTING_CHECKLISTS          # DD sites automated pre-counting

    stage_minutes = {s: by_stage[s] * MINUTES_PER_STAGE[s] for s in by_stage}
    minutes = sum(stage_minutes.values())
    dd_runs = counts.get("single", 0) + counts.get("assemblage", 0)

    breakdown = [{
        "stage": s, "label": STAGE_LABELS[s], "runs": by_stage[s],
        "minutes_per": MINUTES_PER_STAGE[s], "minutes_saved": stage_minutes[s],
        "hours_saved": round(stage_minutes[s] / 60, 1),
    } for s in STAGE_ORDER]

    return {
        "app_runs": dd_runs, "starting": STARTING_CHECKLISTS,
        "total_automated": by_stage["dd"],          # DD checklists incl. baseline (back-compat)
        "minutes_per": MINUTES_PER_CHECKLIST,        # DD per-run minutes (back-compat)
        "minutes_saved": minutes, "hours_saved": round(minutes / 60, 1),
        "total_runs": sum(int(v) for v in counts.values()),
        "by_stage": breakdown,
    }


# --------------------------------------------------------------------------- #
# per-device usage tally (silent attribution)
#
# Each browser/device is tagged with a stable anonymous id (a long-lived cookie
# set by app.py); the connecting IP is recorded as a secondary hint. We keep a
# per-device run count, persisted to the volume so it survives redeploys. This
# is rough usage attribution, NOT identity: laptop+phone = two devices, an office
# NAT collapses people to one IP (the cookie disambiguates), and clearing cookies
# looks like a new device. An admin labels a device with a person's name.
# --------------------------------------------------------------------------- #
def _read_devices():
    try:
        return json.loads(DEVICE_FILE.read_text())
    except Exception:
        return {}


def _write_devices(d):
    try:
        DEVICE_FILE.write_text(json.dumps(d))
    except Exception:
        traceback.print_exc()


def _touch_rec(d, did, ip, now):
    """Upsert a device record's presence fields (first/last seen, IPs). Caller holds _device_lock."""
    rec = d.get(did) or {"first_seen": now, "label": None, "ips": [], "counts": {}, "last_ip": ip}
    rec["last_seen"] = now
    if ip:
        rec["last_ip"] = ip
        if ip not in rec["ips"]:
            rec["ips"] = (rec["ips"] + [ip])[-10:]   # keep the last 10 distinct IPs
    d[did] = rec
    return rec


def touch_device(actor):
    """Register a device's presence without counting a run — called on app load so
    even people who only browse appear in the roster. Best-effort; never raises."""
    if not actor or not actor.get("device"):
        return
    with _device_lock:
        d = _read_devices()
        _touch_rec(d, actor["device"], actor.get("ip") or "", _now())
        _write_devices(d)


def _bump_device(actor, kind):
    """Increment a device's run count for one completed job. Best-effort; never raises."""
    if not actor or not actor.get("device"):
        return
    with _device_lock:
        d = _read_devices()
        rec = _touch_rec(d, actor["device"], actor.get("ip") or "", _now())
        rec["counts"][kind] = int(rec["counts"].get(kind, 0)) + 1
        _write_devices(d)


def device_totals():
    """Per-device usage breakdown for the admin view, newest-active first.

    'dd_runs' = single + assemblage (DD checklists); 'hours_saved' now spans the
    whole pipeline (each kind's runs x its stage minutes), so a person's number is
    comparable to the banner."""
    d = _read_devices()
    out = []
    for did, rec in d.items():
        counts = rec.get("counts", {}) or {}
        dd = int(counts.get("single", 0)) + int(counts.get("assemblage", 0))
        total = sum(int(v) for v in counts.values())
        minutes = sum(int(n) * MINUTES_PER_STAGE.get(STAGE_OF_KIND[k], 0)
                      for k, n in counts.items() if k in STAGE_OF_KIND)
        out.append({
            "device": did, "short": did[:8], "label": rec.get("label"),
            "last_ip": rec.get("last_ip"), "ips": rec.get("ips", []),
            "first_seen": rec.get("first_seen"), "last_seen": rec.get("last_seen"),
            "counts": counts, "total_runs": total, "dd_runs": dd,
            "hours_saved": round(minutes / 60, 1),
        })
    out.sort(key=lambda r: r["last_seen"] or "", reverse=True)
    return out


def set_device_label(device_id, label):
    """Name a device (admin only). Returns False if the device id is unknown."""
    with _device_lock:
        d = _read_devices()
        if device_id not in d:
            return False
        d[device_id]["label"] = (label or "").strip() or None
        _write_devices(d)
        return True


def _job_counts(j):
    """(field count, flag count) — from the live fields dict, else persisted stub counts."""
    with j["_lock"]:
        states = [f["state"] for f in j["fields"].values()]
    if states:
        flags = sum(1 for s in states if s.startswith("TOOL-FAIL") or s == "MANUAL-VERIFY")
        return len(states), flags
    return j.get("_n_fields", 0), j.get("_n_flags", 0)


def _downloadable(j):
    return bool(j.get("file")) and os.path.exists(j["file"])


def recent_jobs(n=12):
    """Most recent completed runs, newest first — re-download + 'generate model' list."""
    with _lock:
        done = [j for j in _jobs.values() if j.get("status") == "done"]
    done.sort(key=lambda j: j.get("finished") or "", reverse=True)
    out = []
    for j in done[:n]:
        nfields, nflags = _job_counts(j)
        out.append({
            "id": j["id"], "kind": j["kind"], "label": j.get("label") or j["id"],
            "finished": j.get("finished"), "fields": nfields, "flags": nflags,
            "downloadable": _downloadable(j),
            # a DD checklist (not a model) with a file on disk can seed a financial model or comp run
            "can_model": j["kind"] in ("single", "assemblage") and _downloadable(j),
            "can_comps": j["kind"] in ("single", "assemblage") and _downloadable(j),
        })
    return out


def _persist_index():
    """Write a compact index of completed jobs to the volume, so the Recent list,
    downloads, and 'generate model' all survive a redeploy (the workbooks themselves
    already persist in DATA_DIR). Best-effort — never raises into a job."""
    with _lock:
        done = [j for j in _jobs.values() if j.get("status") == "done" and _downloadable(j)]
    done.sort(key=lambda j: j.get("finished") or "", reverse=True)
    recs = []
    for j in done[:MAX_JOBS]:
        nfields, nflags = _job_counts(j)
        recs.append({"id": j["id"], "kind": j["kind"], "label": j.get("label") or j["id"],
                     "finished": j.get("finished"), "file": j.get("file"),
                     "filename": j.get("filename"), "n_fields": nfields, "n_flags": nflags,
                     "underwrite": bool(j.get("underwrite"))})
    with _index_lock:
        try:
            INDEX_FILE.write_text(json.dumps(recs))
        except Exception:
            traceback.print_exc()


def _load_index():
    """Rehydrate completed-job stubs from the volume index at startup (in-memory _jobs
    is otherwise empty after a redeploy). Stubs carry enough to list, download, and seed
    a model; their live field detail is gone, so counts come from the persisted index."""
    try:
        recs = json.loads(INDEX_FILE.read_text())
    except Exception:
        return
    for r in recs:
        if r.get("id") in _jobs or not r.get("file"):
            continue
        _jobs[r["id"]] = {
            "id": r["id"], "kind": r.get("kind", "single"), "status": "done",
            "label": r.get("label"), "file": r.get("file"), "filename": r.get("filename"),
            "finished": r.get("finished"), "started": r.get("finished"),
            "fields": {}, "_n_fields": r.get("n_fields", 0), "_n_flags": r.get("n_flags", 0),
            "underwrite": {"models": []} if r.get("underwrite") else None,
            "geo": None, "in_la_city": None, "phase": "Complete",
            "total": r.get("n_fields", 0), "completed": r.get("n_fields", 0),
            "parcels": None, "combined_sf": None, "om": None, "error": None,
            "input": {}, "_lock": threading.Lock(), "_rehydrated": True,
        }


def _prune():
    """Evict oldest jobs (and their files) beyond MAX_JOBS. Caller holds _lock."""
    if len(_jobs) <= MAX_JOBS:
        return
    for jid in sorted(_jobs, key=lambda j: _jobs[j]["started"])[:len(_jobs) - MAX_JOBS]:
        old = _jobs.pop(jid)
        if old.get("file"):
            try:
                Path(old["file"]).unlink(missing_ok=True)
            except Exception:
                pass


def create_job(kind, payload, actor=None):
    jid = uuid.uuid4().hex[:12]
    label = payload.get("address") or ", ".join(payload.get("apns", [])) or jid
    job = {
        "id": jid, "kind": kind, "status": "running", "input": payload, "label": label,
        "geo": None, "in_la_city": None, "phase": "Starting…",
        "total": 0, "completed": 0, "fields": {},
        "parcels": None, "combined_sf": None, "om": None, "underwrite": None,
        "error": None, "file": None, "filename": None,
        "actor": actor or {},          # {device, ip} — silent attribution, set by app.py from the request
        "started": _now(), "finished": None,
        "_lock": threading.Lock(),
    }
    with _lock:
        _jobs[jid] = job
        _prune()
    threading.Thread(target=_run, args=(job,), daemon=True).start()
    return jid


_RUNNERS = {"single": run_single, "assemblage": run_assemblage, "underwrite": run_underwrite,
            "lihtc_scenarios": run_lihtc_scenarios, "pdf_summary": run_pdf_summary,
            "comps": run_comps, "comps_grid": run_comps_grid}


def _run(job):
    try:
        _RUNNERS[job["kind"]](job)
        job["status"] = "done"
        if job["kind"] in STAGE_OF_KIND:
            _bump_count(job["kind"])   # one completed pipeline stage = its MINUTES_PER_STAGE saved
    except Exception as e:
        traceback.print_exc()
        job["status"] = "error"
        job["error"] = str(e)
        job["phase"] = "Error"
    finally:
        job["finished"] = _now()
        if job["status"] == "done":
            _bump_device(job.get("actor"), job["kind"])   # silent per-device attribution
            _persist_index()           # keep the volume index in sync with completed runs


def get_job(jid):
    return _jobs.get(jid)


def public_view(job):
    """Serialize a job for the API (omits the on-disk path + internal lock)."""
    with job["_lock"]:
        fields = list(job["fields"].values())
    return {
        "id": job["id"], "kind": job["kind"], "status": job["status"],
        "phase": job["phase"], "geo": job["geo"], "in_la_city": job["in_la_city"],
        "total": job["total"], "completed": job["completed"],
        "fields": fields,
        "parcels": job["parcels"], "combined_sf": job["combined_sf"],
        "om": job.get("om"), "underwrite": job.get("underwrite"),
        "error": job["error"], "downloadable": _downloadable(job),
        "started": job["started"], "finished": job["finished"],
    }


# Rehydrate prior completed runs from the volume so they're available immediately
# (the Recent list + 'generate model' work right after a redeploy, not just after a run).
_load_index()
