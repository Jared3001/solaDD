"""
tracker_export.py — push completed feasibility runs into the team's shared
Google Sheet ("Acquisition Pipeline" tab of the Main LIHTC Tracker).

Source of truth is the FEASIBILITY PIPELINE (the DD checklist + the LIHTC
scenario one-pager), not the standalone modularZ tool. Two write moments:

  * DD job (single / assemblage) completes  -> descriptive block, cols A-L.
  * One-pager job (pdf_summary) completes    -> the scenario cost blocks for
    whichever construction scenarios ran (5-Story Modular / 12-Story Modular /
    Stick Built).

Rows are upserted by APN OVERLAP so a deal that is re-run — or run first as a
single parcel and later as an assemblage that *includes* that parcel — updates
the SAME row instead of duplicating. The automation owns a hidden "Deal Key"
column (BZ) holding the normalized APN set; it never matches on human-edited
columns (Project Name / Address), which are inconsistent in the live sheet.

Auth — SOLA Workspace blocks external sharing, so we cannot share the sheet
with the service-account email directly. Default path is a service account with
DOMAIN-WIDE DELEGATION impersonating an internal user who already has edit
access (no external share happens). A Workspace Super Admin authorizes the SA's
client id + the spreadsheets scope once. OAuth-user creds are the no-admin
fallback.

Config (environment):
  TRACKER_SHEET_ID              spreadsheet id (required to write)
  TRACKER_TAB                   worksheet name (default "Acquisition Pipeline")
  GOOGLE_SA_JSON                service-account key JSON (string)
  GOOGLE_IMPERSONATE_SUBJECT    internal user email to impersonate (DWD)
  GOOGLE_OAUTH_AUTHORIZED_USER  authorized-user JSON (OAuth fallback)
  TRACKER_DRY_RUN=1             compute + log the row, never write (testing)

With no credential configured the module is a silent no-op, so it is safe to
deploy before the Google side is provisioned.
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path

import yaml
from openpyxl import load_workbook

# --------------------------------------------------------------------------- #
# layout — matches the live "Acquisition Pipeline" tab (header row 4, data 5+)
# --------------------------------------------------------------------------- #
ROOT = Path(__file__).resolve().parent.parent
DD_SHEET = "Site DD"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

HEADER_ROW = 4
FIRST_DATA_ROW = 5
DEAL_KEY_COL = 78          # column BZ — automation-owned dedup key (hidden)
DEAL_KEY_HEADER = "Deal Key (auto — do not edit)"

# Descriptive block: 1-based tracker column -> field name in the desc dict.
# A=1 .. L=12.  Neighborhood (E) is reverse-geocoded from the address.
DESC_COLS = {
    1:  "project_name",
    2:  "units",
    3:  "county",
    4:  "city",
    5:  "neighborhood",
    6:  "geographic_pool",
    7:  "address",
    8:  "qct_dda",
    9:  "resource_area",
    10: "set_aside",          # Non Targeted / Large Family / Senior
    11: "land_price",
    12: "land_sf",
}

# First column (1-based) of each 8-wide scenario block.
#   Low-rise Modular -> M(13);  High-rise Modular -> U(21);  Stick (low-rise) -> AC(29)
# Keyed by (construction, height-tier) rather than a literal story count: stories
# are now analyst-editable in the picker (e.g. a 4-story Stick), but each scenario
# still maps to the same conceptual tracker column. Tier = "high" at 8+ stories.
SCEN_BLOCK_START = {
    ("Modular", "low"):  13,
    ("Modular", "high"): 21,
    ("Stick",   "low"):  29,
    ("Stick",   "high"): 29,  # no separate Stick high-rise column; share the slot
}


def _height_tier(stories):
    try:
        return "high" if int(stories) >= 8 else "low"
    except (TypeError, ValueError):
        return "low"
# Field order within a block (offset 0..7), keyed to _calc()'s results dict.
SCEN_FIELDS = [
    "tdc_unit",          # TDC/Unit
    "tdc",               # TDC
    "cap_dev_fee_30",    # Capitalized Dev Fee 30%
    "cap_dev_fee_70",    # Capitalized Dev Fee 70%
    "deferred_dev_fee",  # Deferred Dev Fee
    "total_dev_fee",     # Total Dev Fee
    "sponsor_funds",     # Sponsor Funds Needed
    "tiebreaker",        # Tiebreaker
]

_POSITIVE = {"yes", "y", "true", "1", "qct", "dda", "qct/dda"}


# --------------------------------------------------------------------------- #
# config / auth
# --------------------------------------------------------------------------- #
def _config():
    """Return write config, or None when the integration is not provisioned."""
    sheet_id = os.environ.get("TRACKER_SHEET_ID")
    if not sheet_id:
        return None
    has_cred = os.environ.get("GOOGLE_SA_JSON") or os.environ.get("GOOGLE_OAUTH_AUTHORIZED_USER")
    if not has_cred and os.environ.get("TRACKER_DRY_RUN") != "1":
        return None
    return {
        "sheet_id": sheet_id,
        "tab": os.environ.get("TRACKER_TAB", "Acquisition Pipeline"),
        "dry_run": os.environ.get("TRACKER_DRY_RUN") == "1",
    }


def _credentials():
    """Build Google credentials from env. DWD service account, else OAuth user."""
    sa = os.environ.get("GOOGLE_SA_JSON")
    if sa:
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_info(
            json.loads(sa), scopes=SCOPES)
        subject = os.environ.get("GOOGLE_IMPERSONATE_SUBJECT")
        return creds.with_subject(subject) if subject else creds
    au = os.environ.get("GOOGLE_OAUTH_AUTHORIZED_USER")
    if au:
        from google.oauth2.credentials import Credentials
        return Credentials.from_authorized_user_info(json.loads(au), SCOPES)
    raise RuntimeError("No Google credential in env.")


def _open_ws(cfg):
    import gspread
    gc = gspread.authorize(_credentials())
    sh = gc.open_by_key(cfg["sheet_id"])
    return sh.worksheet(cfg["tab"])


# --------------------------------------------------------------------------- #
# DD descriptive block
# --------------------------------------------------------------------------- #
def _dd_cell_map():
    schema = yaml.safe_load((ROOT / "canonical" / "schema.yaml").read_text())
    return {f["id"]: f.get("answer_cell") for f in schema["fields"]}


def _project_name(address):
    try:
        import uw_logic
        return uw_logic.project_name(address) or address
    except Exception:
        return address


# Reverse-geocode endpoint for the colloquial neighborhood name. OSM returns
# locality components; for LA the "suburb" field is the broad neighborhood the
# tracker uses (e.g. "Echo Park", "Pico-Union"), with finer fallbacks after it.
_NOMINATIM_REVERSE = "https://nominatim.openstreetmap.org/reverse"
_NEIGHBORHOOD_KEYS = ("suburb", "neighbourhood", "quarter", "city_district",
                      "residential", "borough")


def _single_address(address):
    """Reduce an assemblage address ('addr1; addr2; ... addrN, CITY, ST, ZIP')
    to one geocodable address: the first parcel's street plus the shared
    city/state/zip tail. A plain single address passes through unchanged."""
    addr = str(address)
    if ";" not in addr:
        return addr
    segs = addr.split(";")
    first = segs[0].strip()
    last = segs[-1].strip()
    tail = last.split(",", 1)[1].strip() if "," in last else ""
    return f"{first}, {tail}" if tail else first


def _neighborhood(address):
    """Best-effort neighborhood from a street address: forward-geocode to
    lat/lon (reusing the pipeline's Census/OSM geocoder), then OSM reverse-
    geocode and take the broadest locality component. Returns None on any
    failure so the column is left blank rather than guessed."""
    if not address:
        return None
    try:
        import sys as _sys
        src = str(ROOT / "build" / "sources")
        if src not in _sys.path:
            _sys.path.insert(0, src)
        from geocoder import geocode
        geo = geocode(_single_address(address))
        lat, lon = geo.get("lat"), geo.get("lon")
        if lat is None or lon is None:
            return None
        import urllib.parse
        import urllib.request
        q = urllib.parse.urlencode({"lat": lat, "lon": lon, "format": "json",
                                    "zoom": 14, "addressdetails": 1})
        req = urllib.request.Request(
            f"{_NOMINATIM_REVERSE}?{q}",
            headers={"User-Agent": "solaDD/1.0 (DD automation)"})
        with urllib.request.urlopen(req, timeout=30) as r:
            d = json.load(r)
        addr = d.get("address", {}) if isinstance(d, dict) else {}
        for key in _NEIGHBORHOOD_KEYS:
            v = addr.get(key)
            if v:
                return str(v).strip()
    except Exception:
        return None
    return None


def _qct_dda(qct, dda):
    parts = []
    if str(qct or "").strip().lower() in _POSITIVE or "qct" in str(qct or "").lower():
        parts.append("QCT")
    if str(dda or "").strip().lower() in _POSITIVE or "dda" in str(dda or "").lower():
        parts.append("DDA")
    return "/".join(parts) if parts else "None"


def descriptive_from_dd(dd_path):
    """Read the DD workbook's Site-DD answers into the tracker descriptive block."""
    cells = _dd_cell_map()
    ws = load_workbook(dd_path, data_only=True)[DD_SHEET]

    def g(fid):
        ref = cells.get(fid)
        return ws[ref].value if ref else None

    address = g("address")
    return {
        "project_name":    _project_name(address),
        "units":           g("estimated_unit_count"),
        "county":          g("county"),
        "city":            g("city_jurisdiction"),
        "neighborhood":    _neighborhood(address),   # reverse-geocoded from address
        "geographic_pool": g("geographic_pool"),
        "address":         address,
        "qct_dda":         _qct_dda(g("qct"), g("dda")),
        "resource_area":   g("resource_area"),
        "set_aside":       None,   # filled from scenario LF flag when available
        "land_price":      g("acquisition_price"),
        "land_sf":         g("land_sf"),
        "_apn_raw":        g("apn"),
    }


# --------------------------------------------------------------------------- #
# APN normalization + overlap keys
# --------------------------------------------------------------------------- #
def _apn_set(raw):
    """Normalize a raw APN string/list into a set of digit-only parcel ids."""
    if raw is None:
        return set()
    if isinstance(raw, (list, tuple)):
        items = raw
    else:
        items = re.split(r"[,;/|]+|\s{2,}", str(raw))
    out = set()
    for it in items:
        norm = re.sub(r"[^0-9A-Za-z]", "", str(it)).upper()
        # Real APNs are numeric (sometimes a trailing letter). Require >=6 digits
        # so sentinels like "(unresolved)" and stray fragments never become keys
        # (otherwise two unrelated assemblages could falsely overlap-match).
        if sum(ch.isdigit() for ch in norm) >= 6:
            out.add(norm)
    return out


def _job_apns(job, desc):
    """Best APN set for a job: assemblage parcels, else the DD apn cell."""
    apns = set()
    for p in job.get("parcels", []) or []:
        apns |= _apn_set(p.get("apn"))
    if not apns:
        apns = _apn_set(desc.get("_apn_raw"))
    return apns


def _parse_key(cell):
    return _apn_set(cell)


# --------------------------------------------------------------------------- #
# scenario blocks
# --------------------------------------------------------------------------- #
def scenario_colvals(scenarios):
    """{1-based col: value} for every scenario block we can place, plus the
    set-aside string inferred from any Large-Family scenario."""
    colvals = {}
    set_aside = None
    for scn in scenarios or []:
        start = SCEN_BLOCK_START.get((scn.get("constr"), _height_tier(scn.get("stories"))))
        if not start:
            continue
        res = scn.get("results") or {}
        for off, field in enumerate(SCEN_FIELDS):
            v = res.get(field)
            if v is not None:
                colvals[start + off] = v
        if str(scn.get("lf", "No")).strip().lower() in ("yes", "y", "true"):
            set_aside = "Large Family"
    if set_aside is None and scenarios:
        set_aside = "Non-Target"
    return colvals, set_aside


def descriptive_colvals(desc):
    out = {}
    for col, field in DESC_COLS.items():
        v = desc.get(field)
        if v is not None and v != "":
            out[col] = v
    return out


# --------------------------------------------------------------------------- #
# upsert
# --------------------------------------------------------------------------- #
def _find_row(ws, key_set):
    """Row index of the first existing row whose Deal-Key set overlaps key_set."""
    if not key_set:
        return None
    col = ws.col_values(DEAL_KEY_COL)        # 1-based, includes header rows
    for i, cell in enumerate(col, start=1):
        if i < FIRST_DATA_ROW or not cell:
            continue
        if _parse_key(cell) & key_set:
            return i
    return None


def _next_empty_row(ws):
    keys = ws.col_values(1)                   # project-name column
    row = max(len(keys) + 1, FIRST_DATA_ROW)
    return row


def _a1(row, col):
    import gspread.utils as u
    return u.rowcol_to_a1(row, col)


def _apply(ws, row, colvals, key_set):
    """Write {col: value} into `row`, skip blanks, stamp the Deal-Key column."""
    cells = []
    for col, val in colvals.items():
        c = ws.cell(row, col)
        c.value = val
        cells.append(c)
    kc = ws.cell(row, DEAL_KEY_COL)
    kc.value = ",".join(sorted(key_set)) if key_set else ""
    cells.append(kc)
    # RAW (not USER_ENTERED): a bare number typed into a percent-formatted cell
    # gets scaled ÷100 by Sheets' value parser. The Tiebreaker column is
    # percent-formatted and existing rows store the un-scaled score (e.g. 2.684
    # displays as "268.40%"), so we must write values verbatim.
    ws.update_cells(cells, value_input_option="RAW")


def _ensure_key_header(ws):
    try:
        if not (ws.cell(HEADER_ROW, DEAL_KEY_COL).value or "").strip():
            ws.update_cell(HEADER_ROW, DEAL_KEY_COL, DEAL_KEY_HEADER)
    except Exception:
        pass


def _upsert(colvals, key_set, label=""):
    """Find-or-append by APN overlap and write the given columns."""
    cfg = _config()
    if cfg is None:
        return {"ok": False, "skipped": "not-configured"}
    if cfg["dry_run"] and not (os.environ.get("GOOGLE_SA_JSON") or
                               os.environ.get("GOOGLE_OAUTH_AUTHORIZED_USER")):
        # pure local dry-run: no Google client, just echo what we'd write
        return {"ok": True, "dry_run": True, "label": label,
                "key": sorted(key_set), "colvals": colvals}

    ws = _open_ws(cfg)
    _ensure_key_header(ws)
    row = _find_row(ws, key_set)
    action = "update" if row else "append"
    if not row:
        row = _next_empty_row(ws)
    if cfg["dry_run"]:
        return {"ok": True, "dry_run": True, "action": action, "row": row,
                "label": label, "key": sorted(key_set), "colvals": colvals}
    _apply(ws, row, colvals, key_set)
    return {"ok": True, "action": action, "row": row, "label": label}


# --------------------------------------------------------------------------- #
# public entry points (called best-effort from jobs._run)
# --------------------------------------------------------------------------- #
def export_dd(job):
    """DD job finished -> upsert the descriptive block."""
    if _config() is None:
        return {"ok": False, "skipped": "not-configured"}
    dd_path = job.get("file")
    if not dd_path or not Path(dd_path).exists():
        return {"ok": False, "skipped": "no-dd-file"}
    desc = descriptive_from_dd(dd_path)
    key = _job_apns(job, desc)
    return _upsert(descriptive_colvals(desc), key, label=desc.get("address") or "")


def export_scenarios(job):
    """One-pager job finished -> upsert the scenario cost blocks (+ descriptive)."""
    if _config() is None:
        return {"ok": False, "skipped": "not-configured"}
    scenarios = job.get("scenarios") or []
    if not scenarios:
        return {"ok": False, "skipped": "no-scenarios"}

    colvals, set_aside = scenario_colvals(scenarios)

    # Re-derive the descriptive block + APN key from the source DD when available,
    # so a row created here carries identity even if the DD push was missed.
    key = set()
    dd_path = job.get("dd_file")
    if dd_path and Path(dd_path).exists():
        desc = descriptive_from_dd(dd_path)
        if set_aside is not None:
            desc["set_aside"] = set_aside
        colvals.update(descriptive_colvals(desc))
        key = _job_apns(job, desc)

    return _upsert(colvals, key, label=job.get("label") or "")
