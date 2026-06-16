#!/usr/bin/env python3
"""
assemblage.py — multi-APN block-assemblage support.

A site can be one legal lot or a BLOCK ASSEMBLAGE of several APNs bought together
(e.g. 4201 Pico = "entire block"). collect.py sizes only the anchor parcel; this
takes a LIST of APNs, sizes the combined site, and aggregates every Tier-A
designation across the parcels — reporting a single site-level answer and
flagging where parcels DIFFER (mixed zoning, a hazard on only some lots, a tract
split, etc.).

Land area: the LA City Parcels layer (Landbase_Information/MapServer/5, keyed by
BPP) reproduces ZIMAS lot areas as polygon geometry in EPSG:2229 (US survey ft).
One APN can map to several lot polygons; we sum them per APN and across the site.

Each parcel runs the same readers as collect.py (statewide via the parcel's
census tract from geocode_point; ZIMAS via the parcel centroid). Aggregation:
  - yes/no hazard fields  -> "Yes" if ANY parcel is Yes (names the APNs); else No
  - text/enum fields      -> the shared value, or "MIXED: v (apns); w (apns)"
The combined land_sf, the APN list, and every aggregated field are written to a
site workbook (cols A/C/E), with per-parcel detail in Notes.

Usage:
  python build/assemblage.py <wb.xlsx> APN1 APN2 ...
  python build/assemblage.py --demo APN1 APN2 ...   (throwaway workbook)
"""
import re
import sys
import shutil
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build"))
sys.path.insert(0, str(ROOT / "build" / "sources"))

import yaml
import _arcgis as ag
from geocoder import geocode_point
from runner import run_reader
# SD_READERS is being added to collect.py concurrently by another agent; import it
# defensively so this tool doesn't crash if it isn't present yet.
try:
    from collect import READERS, ZIMAS_READERS, SD_READERS
except ImportError:
    from collect import READERS, ZIMAS_READERS
    SD_READERS = {}
import zimas, nc
import sd_parcel

LACITY_PARCELS = "https://maps.lacity.org/lahub/rest/services/Landbase_Information/MapServer"
PARCEL_LAYER = 5   # "Parcels" — geometry area in EPSG:2229 reproduces ZIMAS lot area

# Fields assemblage computes itself at the site level (not per-parcel aggregated):
# combined land area + the per-APN list (address is a per-parcel placeholder here).
_SITE_LEVEL = {"land_sf", "apn", "address"}


def _norm(apn: str) -> str:
    """'5082-024-018' -> '5082024018' (BPP key)."""
    return re.sub(r"[^0-9A-Za-z]", "", apn).upper()


def _ring_area_sf(rings):
    """Shoelace area (sq ft) of an ArcGIS polygon whose coords are in EPSG:2229 ft."""
    tot = 0.0
    for ring in rings:
        s = 0.0
        for i in range(len(ring) - 1):
            s += ring[i][0] * ring[i + 1][1] - ring[i + 1][0] * ring[i][1]
        tot += s / 2.0
    return abs(tot)


def _parcel_info_la(apn: str) -> dict:
    """Resolve an APN against LA City Parcels (BPP, EPSG:2229). Returns the parcel
    dict (with jurisdiction='LA') or raises LookupError if the APN isn't there."""
    bpp = _norm(apn)
    where = f"BPP='{bpp}'"
    feet = ag.query(LACITY_PARCELS, PARCEL_LAYER, where=where, return_geometry=True,
                    out_sr=2229, out_fields="BPP,LOT")
    if not feet:
        raise LookupError(f"no LA City parcel for APN {apn} (BPP {bpp})")
    land_sf = sum(_ring_area_sf(f["geometry"]["rings"]) for f in feet if f.get("geometry"))
    wgs = ag.query(LACITY_PARCELS, PARCEL_LAYER, where=where, return_geometry=True,
                   out_sr=4326, out_fields="BPP")
    # representative point = centroid of the largest lot polygon (in lon/lat)
    best_pt, best_a = None, -1.0
    for f in wgs:
        rings = (f.get("geometry") or {}).get("rings") or []
        if not rings:
            continue
        pts = rings[0]
        a = len(pts)
        if a > best_a:
            best_a = a
            best_pt = (sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts))
    lon, lat = best_pt
    geo = geocode_point(lon, lat)
    return {"apn": apn, "jurisdiction": "LA", "bpp": bpp, "n_lots": len(feet),
            "land_sf": round(land_sf, 1),
            "lon": lon, "lat": lat, "geoid": geo["geoid"],
            "state_fips": geo["state_fips"], "county_fips": geo["county_fips"],
            "place": geo.get("place"), "matched_address": geo["matched_address"]}


def _parcel_info_sd(apn: str) -> dict:
    """Resolve an APN against SANDAG County Parcels (EPSG:2230). sd_parcel returns
    {apn, n_lots, land_sf, lon, lat} but no tract/FIPS, so we fill geoid/state_fips/
    county_fips/place/matched_address from geocode_point on the centroid — the same
    geo dict the statewide readers consume. Returns dict with jurisdiction='SD'."""
    info = sd_parcel.parcel_info(apn)   # raises LookupError if not in SANDAG
    lon, lat = info["lon"], info["lat"]
    if lon is None or lat is None:
        raise LookupError(f"SANDAG parcel {apn} has no centroid for geo resolution")
    geo = geocode_point(lon, lat)
    return {"apn": info["apn"], "jurisdiction": "SD", "bpp": _norm(apn),
            "n_lots": info["n_lots"], "land_sf": info["land_sf"],
            "lon": lon, "lat": lat, "geoid": geo["geoid"],
            "state_fips": geo["state_fips"], "county_fips": geo["county_fips"],
            "place": geo.get("place"), "matched_address": geo["matched_address"]}


def parcel_info(apn: str) -> dict:
    """Resolve an APN to a parcel dict, auto-detecting jurisdiction.

    LA-City and San Diego APNs are both 10 digits (just different dash grouping:
    '5082-024-018' vs '453-122-10-00'), so length can't disambiguate them. Instead
    we detect by which parcel layer actually resolves the APN: try LA City Parcels
    (BPP) first, and if that returns nothing, try SANDAG. The result carries a
    'jurisdiction' marker ('LA'/'SD') so the reader aggregation can pick the right
    municipal block (ZIMAS for LA, SD_READERS for San Diego)."""
    try:
        return _parcel_info_la(apn)
    except LookupError:
        pass
    try:
        return _parcel_info_sd(apn)
    except LookupError:
        pass
    raise LookupError(
        f"APN {apn} not found in LA City Parcels or SANDAG County Parcels")


def _parcel_geo(parcel):
    return {"lon": parcel["lon"], "lat": parcel["lat"], "geoid": parcel["geoid"],
            "state_fips": parcel["state_fips"], "county_fips": parcel["county_fips"],
            "place": parcel.get("place"), "matched_address": parcel["matched_address"]}


def _run_all_parcels(parcels, workers=12):
    """Run every (parcel, reader) task through one thread pool; return per[apn][fid].

    Snaps are pre-warmed per parcel (different cache keys, no race) and the
    Neighborhood-Change file once, so the parallel fan-out only does fresh reads.

    Statewide READERS run on every parcel; the municipal block is jurisdiction-
    specific — ZIMAS for LA parcels, SD_READERS for San Diego parcels. (Mixed-
    jurisdiction assemblages are unusual but handled per-parcel here.)
    """
    base = {k: v for k, v in READERS.items() if k not in _SITE_LEVEL}
    muni = {
        "LA": {k: v for k, v in ZIMAS_READERS.items() if k not in _SITE_LEVEL},
        "SD": {k: v for k, v in SD_READERS.items() if k not in _SITE_LEVEL},
    }
    try:
        nc._load()
    except Exception:
        pass
    # Only LA parcels share the ZIMAS snap cache; SD readers carry their own snap.
    la_parcels = [p for p in parcels if p.get("jurisdiction") == "LA"]
    if la_parcels:
        with ThreadPoolExecutor(max_workers=min(6, len(la_parcels))) as ex:
            list(ex.map(lambda p: _warm(_parcel_geo(p)), la_parcels))

    tasks = [(p, fid, fn) for p in parcels
             for fid, fn in {**base, **muni.get(p.get("jurisdiction"), {})}.items()]

    def _task(t):
        p, fid, fn = t
        geo = _parcel_geo(p)
        kind, payload = run_reader(lambda: fn(geo))
        if kind == "ok":
            return p["apn"], fid, {"answer": payload.get("answer"), "state": payload.get("state", "VERIFIED")}
        return p["apn"], fid, {"answer": None, "state": "TOOL-FAIL", "error": str(payload)[:120]}

    with ThreadPoolExecutor(max_workers=workers) as ex:
        raw = list(ex.map(_task, tasks))
    per = {"_order": parcels}
    for p in parcels:
        per[p["apn"]] = {}
    for apn, fid, res in raw:
        per[apn][fid] = res
    return per


def _warm(geo):
    try:
        zimas.in_la_city(geo)   # warms this parcel's snap cache
    except Exception:
        pass


def _aggregate(fid, atype, per_parcel):
    """Collapse per-parcel answers into one site answer + note + state."""
    # A municipal field (ZIMAS / SD) only runs on parcels of that jurisdiction, so
    # only aggregate over parcels that actually produced this field.
    vals = {p["apn"]: per_parcel[p["apn"]][fid]
            for p in per_parcel["_order"] if fid in per_parcel[p["apn"]]}
    if not vals:
        return {"answer": None, "state": "MANUAL",
                "notes": "no parcel ran this field (jurisdiction has no reader for it)"}
    fails = [a for a, v in vals.items() if v["state"] == "TOOL-FAIL"]
    ok = {a: v for a, v in vals.items() if v["state"] != "TOOL-FAIL"}
    if not ok:
        return {"answer": None, "state": "TOOL-FAIL",
                "notes": f"all parcels failed: {'; '.join(fails)}"}

    def grp(pred):  # apns whose answer matches pred
        return [a for a, v in ok.items() if pred(v["answer"])]

    if atype == "yes_no":
        yes = grp(lambda x: str(x).strip().lower().startswith("y"))
        ans = "Yes" if yes else "No"
        note = (f"Yes on APN(s): {', '.join(yes)}" if yes
                else f"No across all {len(ok)} parcel(s)")
    else:  # text / enum / number — distinct values
        by_val = {}
        for a, v in ok.items():
            by_val.setdefault(str(v["answer"]), []).append(a)
        if len(by_val) == 1:
            ans = next(iter(by_val))
            note = f"uniform across {len(ok)} parcel(s)"
        else:
            ans = "MIXED — see notes"
            note = "MIXED: " + "; ".join(f"{val} ({', '.join(ap)})" for val, ap in by_val.items())
    # state: any borderline parcel makes the whole site a JUDGMENT (route up);
    # otherwise a single shared state propagates, else VERIFIED.
    states = {v["state"] for v in ok.values()}
    if "JUDGMENT" in states:
        state = "JUDGMENT"
    elif len(states) == 1:
        state = states.pop()
    else:
        state = "VERIFIED"
    if fails:
        note += f" | NOTE: {len(fails)} parcel(s) failed read: {', '.join(fails)}"
    return {"answer": ans, "state": state, "notes": note}


def assess(wb_path, apns, property_id=None):
    schema = yaml.safe_load((ROOT / "canonical/schema.yaml").read_text())
    by_id = {f["id"]: f for f in schema["fields"]}
    atype = {f["id"]: f.get("answer_type") for f in schema["fields"]}

    print(f"resolving {len(apns)} APN(s)…")
    with ThreadPoolExecutor(max_workers=min(6, len(apns))) as ex:
        parcels = list(ex.map(parcel_info, apns))   # map preserves input order
    for p in parcels:
        print(f"  {p['apn']:16} BPP {p['bpp']:12} {p['n_lots']} lot(s)  {p['land_sf']:>10,.1f} sf  tract {p['geoid']}")
    combined = round(sum(p["land_sf"] for p in parcels), 1)
    print(f"  COMBINED LAND AREA: {combined:,.1f} sf  ({combined/43560:.3f} ac)\n")

    per = _run_all_parcels(parcels)

    # write to workbook
    from openpyxl import load_workbook
    wb = load_workbook(wb_path)
    ws = wb["Site DD"]

    def put(cell, answer, state, notes):
        r = int(cell[1:])
        ws.cell(r, 1, state)
        ws.cell(r, 3, answer)
        ws.cell(r, 5, notes)

    apn_list = ", ".join(p["apn"] for p in parcels)
    put(by_id["apn"]["answer_cell"], apn_list, "VERIFIED",
        "Block assemblage: " + "; ".join(f"{p['apn']} ({p['n_lots']} lot(s), {p['land_sf']:,.0f} sf)" for p in parcels))
    put(by_id["land_sf"]["answer_cell"], combined, "VERIFIED",
        f"Combined gross land area across {len(parcels)} APN(s) = {combined:,.1f} sf "
        f"({combined/43560:.3f} ac). Per-APN: " +
        "; ".join(f"{p['apn']}={p['land_sf']:,.0f}" for p in parcels) +
        ". Source: LA City Parcels geometry (EPSG:2229).")

    print(f"{'field':30} {'SITE answer':24} {'state':10} detail")
    print("-" * 100)
    agg = {}
    # Union of every reader that could have run across the assemblage's parcels:
    # statewide + whichever municipal blocks are present (ZIMAS for LA, SD for SD).
    have_la = any(p.get("jurisdiction") == "LA" for p in parcels)
    have_sd = any(p.get("jurisdiction") == "SD" for p in parcels)
    fids = list(READERS)
    if have_la:
        fids += list(ZIMAS_READERS)
    if have_sd:
        fids += list(SD_READERS)
    seen = set()
    for fid in fids:
        if fid in seen:
            continue
        seen.add(fid)
        if fid not in by_id or fid in _SITE_LEVEL:   # site-level fields written above
            continue
        a = _aggregate(fid, atype.get(fid), per)
        agg[fid] = a
        put(by_id[fid]["answer_cell"], a["answer"], a["state"], a["notes"])
        print(f"{fid:30} {str(a['answer'])[:24]:24} {a['state']:10} {a['notes'][:48]}")

    wb.save(wb_path)
    print(f"\nworkbook: {wb_path}")
    return {"parcels": parcels, "combined_sf": combined, "fields": agg}


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--demo":
        apns = args[1:]
        tmp = ROOT / "build" / "_assemblage_demo.xlsx"
        shutil.copy(ROOT / "template/Checklist_BLANK_master.xlsx", tmp)
        assess(tmp, apns, property_id="ASSEMBLAGE")
    elif len(args) >= 2:
        assess(args[0], args[1:])
    else:
        print(__doc__)
