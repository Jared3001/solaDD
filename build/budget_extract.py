#!/usr/bin/env python3
"""
budget_extract.py — normalize a SoLa internal underwriting workbook into one record.

SoLa keeps a STANDARDIZED internal underwriting workbook per asset (the
"<deal> Budget.xlsx" files). The same template repeats across the ~28 portfolio
properties on the workbook's `Property List` tab. This extractor reads the tabs
that carry the reusable signal and emits one normalized JSON record per asset, so
a portfolio of these can be assembled into a calibration corpus for market-rate
modeling (cost / financing / return-schema signal; rents are calibration-only
because most SoLa deals are restricted — see the `restricted` flag).

Tabs read:
  * Budget Summary / Mgmt Summary  -> deal summary: unit mix, TDC, NOI, returns,
                                      debt terms, schedule (label/value grids).
  * GMAX                           -> GC contract totals (GMAX vs current).
  * Lender Budget                  -> hard-cost detail by CSI division.
  * Rent Roll                      -> unit-level market + actual rents, occupancy.

DESIGN: the summary tabs are label/value GRIDS, not clean tables, and cell
positions drift between the Budget Summary and Mgmt Summary layouts (C/D vs B/C
pairs; G/H/I vs F/G/H metric blocks). So extraction is LABEL-DRIVEN: find a label
cell by substring, then read the non-empty values to its right. Nothing is read by
hard-coded cell address. Every field that can't be found is recorded in `flags`
rather than silently dropped.

Usage:
  .venv/bin/python3 build/budget_extract.py "<file.xlsx>" [--out record.json] [--pretty]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# low-level cell helpers (label-driven, position-tolerant)
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    """Lower-cased, whitespace-collapsed string for fuzzy label matching."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def find_label(ws, *substrings, max_row=60, max_col=15):
    """Return (row, col) of the first cell whose text contains ALL substrings."""
    wants = [_norm(s) for s in substrings]
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, min(ws.max_column, max_col) + 1):
            cell = _norm(ws.cell(r, c).value)
            if cell and all(w in cell for w in wants):
                return (r, c)
    return None


def values_right(ws, loc, n=3):
    """Non-empty cell values to the right of a label cell, in column order.

    Bounded by `n` columns and stops at the first TEXT cell encountered after a
    numeric value has been collected — the summary tabs place a second label/value
    block (e.g. PROJECT STATUS) a few columns to the right of each metric block, and
    an unbounded scan bleeds into it. Empty cells are skipped (the 'IC | Projected'
    blocks leave the IC column blank for some metrics, e.g. Stab. NOI).
    """
    if not loc:
        return []
    r, c = loc
    out = []
    for cc in range(c + 1, c + 1 + n):
        if cc > ws.max_column:
            break
        v = ws.cell(r, cc).value
        if v is None or _norm(v) == "":
            continue
        if num(v) is None and any(num(x) is not None for x in out):
            break  # next block's label — stop before bleeding into it
        out.append(v)
    return out


def value_below(ws, *subs):
    """Value in the cell directly below a label (header-over-value layout)."""
    loc = find_label(ws, *subs)
    if not loc:
        return None
    r, c = loc
    return ws.cell(r + 1, c).value


def label_first(ws, *subs, **kw):
    """First value to the right of a label (e.g. left-block 'label: value' pairs)."""
    vals = values_right(ws, find_label(ws, *subs), **kw)
    return vals[0] if vals else None


def label_last(ws, *subs, **kw):
    """Last value to the right of a label.

    The summary metric blocks lay out 'IC | Projected' (or 'IC | Current'); the
    rightmost value is the projected/current figure, which is the one we want.
    """
    vals = values_right(ws, find_label(ws, *subs), **kw)
    return vals[-1] if vals else None


# ---------------------------------------------------------------------------
# typed coercion
# ---------------------------------------------------------------------------

def num(v):
    """Coerce to float; strip $ , % and parens. Return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()%")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    val = float(m.group())
    return -val if neg else val


def isodate(v):
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    return None


# ---------------------------------------------------------------------------
# unit-mix parsing  e.g. "6 (21%) 274 SF"  /  "22 (79%) 362 SF"
# ---------------------------------------------------------------------------

_MIX_RE = re.compile(r"(\d+)\s*\((\d+)\s*%\)\s*(\d+)\s*SF", re.I)


def parse_mix_cell(label, value):
    m = _MIX_RE.search(str(value or ""))
    if not m:
        return None
    return {
        "type": str(label).replace(":", "").strip(),
        "count": int(m.group(1)),
        "pct": round(int(m.group(2)) / 100, 4),
        "avg_sf": int(m.group(3)),
    }


def extract_unit_mix(ws):
    """Unit-mix lines sit next to 'Unit Mix'/'Studios'/'One/Two/Three Beds'."""
    mix = []
    for key in ("studio", "one bed", "two bed", "three bed", "four bed"):
        loc = find_label(ws, key)
        if not loc:
            continue
        r, c = loc
        label = ws.cell(r, c).value
        for v in values_right(ws, loc, n=4):
            row = parse_mix_cell(label, v)
            if row:
                mix.append(row)
                break
    # de-dupe by type, keep first
    seen, out = set(), []
    for m in mix:
        if m["type"].lower() not in seen:
            seen.add(m["type"].lower())
            out.append(m)
    return out


# ---------------------------------------------------------------------------
# section extractors
# ---------------------------------------------------------------------------

def pick_summary_sheet(wb):
    """Prefer 'Budget Summary' (full S&U); fall back to 'Mgmt Summary'."""
    for name in ("Budget Summary", "Mgmt Summary"):
        if name in wb.sheetnames:
            return wb[name]
    return None


def extract_property(ws, flags):
    units = num(label_first(ws, "# of units")) or num(label_last(ws, "# of units"))
    aff = label_first(ws, "affordability")
    aff_s = _norm(aff)
    # restricted if affordability mentions AMI / a restricted %, or dev type says so
    devtype = label_first(ws, "development type")
    restricted = bool(
        ("ami" in aff_s) or ("%" in str(aff or "") and "0" not in aff_s[:1])
        or ("restrict" in _norm(devtype)) or ("section 8" in _norm(devtype))
    )
    prop = {
        "name": label_first(ws, "project name"),
        "fund": label_first(ws, "sola fund"),
        "address": label_first(ws, "address"),
        "city_state_zip": label_first(ws, "city, state"),
        "submarket": label_first(ws, "submarket"),
        "project_type": label_first(ws, "type of project"),
        "product_type": label_first(ws, "product type"),
        "development_type": devtype,
        "zoning": label_first(ws, "zoning"),
        "council_district": label_first(ws, "council district"),
        "units": units,
        "avg_unit_size_sf": num(label_first(ws, "avg unit size")),
        "lot_sf": num(label_first(ws, "lot area")),
        # "Lot Acre / Density" puts acres then density on one row -> density is the larger
        "density_u_per_acre": max(
            [v for v in (num(x) for x in values_right(ws, find_label(ws, "density"))) if v is not None],
            default=None,
        ),
        "gross_bldg_sf": num(value_below(ws, "gross bldg")),
        "net_rentable_sf": num(value_below(ws, "net rent")),
        "building_efficiency": num(value_below(ws, "building efficiency")),
        "affordability": aff,
        "restricted": restricted,
        "unit_mix": extract_unit_mix(ws),
    }
    if not prop["units"]:
        flags.append("property.units not found")
    return prop


def extract_returns(ws):
    return {
        "stab_noi": num(label_last(ws, "stab", "noi")) or num(label_last(ws, "stabilized noi")),
        "stab_coc": num(label_last(ws, "stabilized coc")),
        "project_irr": num(label_last(ws, "project irr")),
        "moic": num(label_last(ws, "moic")),
        "roc_trended": num(label_last(ws, "roc")),
        "avg_rent_psf": num(label_last(ws, "avg rent (psf)")),
        "avg_rent_actual": num(label_last(ws, "avg rent (actual)")),
        "tdc_total": num(label_last(ws, "total dev cost")),
        "tdc_per_unit": num(label_last(ws, "tdc/unit")),
        "land_price_per_unit": num(label_last(ws, "land price/unit")),
    }


def extract_financing(ws):
    fin = {
        "equity_pct": None, "equity_amount": None,
        "constr_loan_pct": None, "constr_loan_amount": None,
        "constr_rate": None, "appraised_value": None,
        "perm_loan_amount": None, "perm_rate": None,
    }
    # sources block: "Equity" / "Construction Loan" rows carry pct then $ amounts
    eq = values_right(ws, find_label(ws, "equity"), n=4)
    if eq:
        nums = [num(x) for x in eq if num(x) is not None]
        if nums:
            fin["equity_pct"] = next((n for n in nums if 0 < n <= 1), None)
            amts = [n for n in nums if n > 1000]
            fin["equity_amount"] = amts[0] if amts else None
    cl = values_right(ws, find_label(ws, "construction loan"), n=4)
    if cl:
        nums = [num(x) for x in cl if num(x) is not None]
        fin["constr_loan_pct"] = next((n for n in nums if 0 < n <= 1), None)
        amts = [n for n in nums if n > 1000]
        fin["constr_loan_amount"] = amts[0] if amts else None
    # debt rates / appraised value mostly live on Mgmt Summary (see extract_debt_terms)
    fin["appraised_value"] = num(label_last(ws, "est appr"))
    return fin


def extract_debt_terms(wb, fin, flags):
    """Mgmt Summary carries the DEBT-CONSTRUCTION / DEBT-PERMANENT blocks."""
    if "Mgmt Summary" not in wb.sheetnames:
        return
    ws = wb["Mgmt Summary"]
    # The DEBT-CONSTRUCTION / DEBT-PERMANENT blocks live in the right-hand columns
    # (J-L) as a loose, partly free-text grid. Rather than chase each label, scan that
    # column band for floats in the mortgage-rate band — LTC/LTV are 0/blank at this
    # stage so they don't collide. First occurrence ~ construction, last ~ permanent.
    rates = []
    for r in range(28, min(ws.max_row, 52) + 1):
        for c in range(10, min(ws.max_column, 12) + 1):
            nv = num(ws.cell(r, c).value)
            if nv is not None and 0.02 <= nv <= 0.09:
                rates.append(nv)
    if rates:
        fin["constr_rate"] = rates[0]
        fin["perm_rate"] = rates[-1]
    else:
        flags.append("financing.rates not found (debt block free-text)")
    appr = num(value_below(ws, "est appr"))
    if appr:
        fin["appraised_value"] = appr
    # perm loan amount is usually free-text ("Perm Debt (At IC): $4M"); parse $ figure
    perm_loc = find_label(ws, "perm debt", max_row=52, max_col=12)
    if perm_loc:
        txt = str(ws.cell(*perm_loc).value)
        m = re.search(r"\$?\s*([\d.]+)\s*M", txt, re.I)
        if m:
            fin["perm_loan_amount"] = float(m.group(1)) * 1_000_000
    if fin["perm_loan_amount"] is None:
        flags.append("financing.perm_loan_amount not found")


def extract_gc_contract(wb, flags):
    gc = {"gmax_contract": None, "current_hc": None, "by_division": {}}
    if "GMAX" in wb.sheetnames:
        ws = wb["GMAX"]
        # the 'X Total' grand-total row: col E = contract amt, col H = total contract
        for r in range(1, ws.max_row + 1):
            a = _norm(ws.cell(r, 1).value)
            if a.endswith("total") and "div" not in a:
                gc["gmax_contract"] = num(ws.cell(r, 5).value)
                gc["current_hc"] = num(ws.cell(r, 8).value)
    if "Lender Budget" in wb.sheetnames:
        ws = wb["Lender Budget"]
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a and _norm(a).startswith("total div"):
                key = re.sub(r"^total\s+", "", str(a).strip(), flags=re.I).strip()
                gc["by_division"][key] = num(ws.cell(r, 2).value)
    if not gc["by_division"]:
        flags.append("cost.by_division empty (no Lender Budget DIV totals)")
    return gc


def extract_rent_roll(wb, flags):
    if "Rent Roll" not in wb.sheetnames:
        flags.append("Rent Roll tab missing")
        return None
    ws = wb["Rent Roll"]
    # locate header row: contains 'Unit Type' and a 'Market' column
    hdr = None
    for r in range(1, min(ws.max_row, 12) + 1):
        rowvals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("unit type" in v for v in rowvals) and any("market" in v for v in rowvals):
            hdr = r
            break
    if hdr is None:
        flags.append("Rent Roll header row not found")
        return None

    def col_of(*subs):
        for c in range(1, ws.max_column + 1):
            cell = _norm(ws.cell(hdr, c).value)
            if cell and all(s in cell for s in subs):
                return c
        return None

    c_unit = col_of("unit") or 1
    c_type = col_of("unit type") or 2
    c_sf = col_of("sq") or col_of("unit", "ft") or 3
    c_mkt = col_of("market") or 6
    c_act = col_of("actual") or 7

    units, by_type = [], {}
    for r in range(hdr + 1, ws.max_row + 1):
        utype = ws.cell(r, c_type).value
        mkt = num(ws.cell(r, c_mkt).value)
        if utype is None or mkt is None:
            # stop at first summary/blank block after we have rows
            if units:
                break
            continue
        rec = {
            "unit": ws.cell(r, c_unit).value,
            "type": str(utype).strip(),
            "sf": num(ws.cell(r, c_sf).value),
            "market_rent": mkt,
            "actual_rent": num(ws.cell(r, c_act).value),
        }
        units.append(rec)
        t = by_type.setdefault(rec["type"], {"count": 0, "sf": [], "market": [], "actual": []})
        t["count"] += 1
        if rec["sf"]:
            t["sf"].append(rec["sf"])
        t["market"].append(rec["market_rent"])
        if rec["actual_rent"] is not None:
            t["actual"].append(rec["actual_rent"])

    def avg(xs):
        return round(sum(xs) / len(xs), 2) if xs else None

    rollup = {
        t: {
            "count": d["count"],
            "avg_sf": avg(d["sf"]),
            "avg_market_rent": avg(d["market"]),
            "avg_actual_rent": avg(d["actual"]),
        }
        for t, d in by_type.items()
    }
    return {
        "as_of": label_first(ws, "as of") or (str(ws.cell(3, 1).value) if ws.cell(3, 1).value else None),
        "unit_count": len(units),
        "total_market_rent": round(sum(u["market_rent"] for u in units), 2),
        "total_actual_rent": round(sum(u["actual_rent"] for u in units if u["actual_rent"]), 2),
        "by_unit_type": rollup,
        "units": units,
    }


def extract_schedule(ws):
    return {
        "close_of_escrow": isodate(label_last(ws, "close of escrow")),
        "constr_start": isodate(label_last(ws, "constr start")),
        "expected_tco": isodate(label_last(ws, "expected tco")),
        "fully_stab_date": isodate(label_last(ws, "fully stab")),
    }


# ---------------------------------------------------------------------------
# orchestrator
# ---------------------------------------------------------------------------

def extract(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    flags: list[str] = []
    summ = pick_summary_sheet(wb)
    if summ is None:
        flags.append("no Budget Summary / Mgmt Summary tab")

    rec = {
        "source_file": path.name,
        "summary_tab": summ.title if summ else None,
        "property": extract_property(summ, flags) if summ else {},
        "returns": extract_returns(summ) if summ else {},
        "financing": extract_financing(summ) if summ else {},
        "cost": extract_gc_contract(wb, flags),
        "schedule": extract_schedule(summ) if summ else {},
        "rent_roll": extract_rent_roll(wb, flags),
        "flags": flags,
    }
    if summ:
        extract_debt_terms(wb, rec["financing"], flags)
        # fold the summary cost metrics into cost
        rec["cost"]["tdc_total"] = rec["returns"].get("tdc_total")
        rec["cost"]["tdc_per_unit"] = rec["returns"].get("tdc_per_unit")
        rec["cost"]["const_cost_psf_net"] = num(value_below(wb["Mgmt Summary"], "const cost psf")) \
            if "Mgmt Summary" in wb.sheetnames else None
    wb.close()
    return rec


def main(argv=None):
    ap = argparse.ArgumentParser(description="Normalize a SoLa underwriting workbook to JSON.")
    ap.add_argument("xlsx", help="path to the budget/underwriting workbook")
    ap.add_argument("--out", help="write JSON here (default: stdout)")
    ap.add_argument("--pretty", action="store_true", help="indent JSON")
    args = ap.parse_args(argv)

    path = Path(args.xlsx).expanduser()
    if not path.exists():
        print(f"error: {path} not found", file=sys.stderr)
        return 2

    rec = extract(path)
    text = json.dumps(rec, indent=2 if args.pretty else None, default=str)
    if args.out:
        Path(args.out).write_text(text)
        n_flags = len(rec["flags"])
        print(f"wrote {args.out}  ({n_flags} flag(s))")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
