#!/usr/bin/env python3
"""
comps.py — comp-collection stage (DD checklist -> financial model -> COMPS -> 1-pager).

Given a subject address + the bed counts to comp, this pulls nearby rental
listings, rolls the unit-level listings up into buildings, shortlists the closest
N as candidate comps, computes the Tier-A math (distance, $/SF, weighted
averages), and writes a rent-comparability grid that mirrors the SoLa CTCAC
template (template/comp_grid_reference.xlsx) — one sheet per bed count.

WHAT THIS FILLS (and the honest limits):
  * Tier A (VERIFIED-from-data): address, city, distance, unit size SF, base rent,
    $/SF, # beds, # baths, year built, plus the computed aggregate/weighted rows.
  * Tier B (JUDGMENT): # units & # stories are ESTIMATED from the listing rollup;
    amenities are filled only if the source carries them.
  * Tier C (MANUAL): the per-row $ ADJUSTMENT columns, vacancy, turnover, waiting
    list, concessions, utilities split — appraiser judgment, left blank on purpose.

Comp SELECTION stays human-in-the-loop: this auto-shortlists candidates; the
analyst confirms which go in the final grid (mirrors the rest of the DD pipeline).

Sources (--source flag):
  ai        Zillow via Firecrawl + Gemini (DEFAULT) — scrapes Zillow rental search
            pages, extracts structured records with Gemini 2.5 Flash (same key as OM
            extraction). Richer amenity data when cards show it; year_built and exact
            sqft sometimes absent from search cards. Needs FIRECRAWL_API_KEY +
            GEMINI_API_KEY. ToS: deal-level use only.
  rentcast  RentCast API — unit-level listings, 50 calls/mo free.
            Needs RENTCAST_API_KEY.

Usage:
  python3 build/comps.py "17719 Kinzie St, Northridge, CA" --beds 1 --demo
  python3 build/comps.py "17719 Kinzie St, Northridge, CA" --beds 0 1 2 --out comps.xlsx
  python3 build/comps.py "17719 Kinzie St, Northridge, CA" --beds 1 --source ai
  (--demo uses the offline fixture for whichever source is selected)
"""
import argparse
import os
import re
import sys
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "sources"))
import _arcgis as ag          # haversine_m(lon1,lat1,lon2,lat2) -> meters
import rentcast
import ai_scraper
from geocoder import geocode

_ZIP_RE = re.compile(r"\b(\d{5})\b")


def _extract_zip(matched_address: str) -> str:
    """Pull the ZIP code out of the Census geocoder's matchedAddress string.
    matchedAddress format: '17719 KINZIE ST, NORTHRIDGE, CA, 91325'
    Take the LAST 5-digit token — the house number at the front would match first."""
    matches = _ZIP_RE.findall(matched_address or "")
    return matches[-1] if matches else ""

M_PER_MILE = 1609.344
BEDS_TO_SHEET = {0: "STUDIO MARKET", 1: "1 BEDROOM MARKET", 2: "2 BEDROOM MARKET",
                 3: "3 BEDROOM MARKET", 4: "4 BEDROOM MARKET"}
_UNIT_RE = re.compile(r"\b(apt|unit|ste|suite|#|no\.?)\s*\S*$", re.I)


def _building_key(rec):
    """Strip the unit token so listings in one building roll up together."""
    addr = (rec.get("address") or "").strip()
    addr = _UNIT_RE.sub("", addr).strip(" ,#")
    return (addr.lower(), rec.get("zip"))


def _avg(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


# Plausible apartment sqft per bedroom count — RentCast carries junk sqft on some
# records (e.g. a "studio" at 113 sf), which would wreck the $/SF value ratio.
# Outside the band -> sqft/$ per SF are dropped (not used in the math) and flagged.
SQFT_BANDS = {0: (300, 750), 1: (450, 1050), 2: (650, 1500), 3: (900, 2000), 4: (1100, 2600)}
_SQFT_FALLBACK = (250, 4000)


def _clean_sqft(units, bedrooms):
    """-> (avg plausible sqft | None, flag_note | ''). Averages only in-band sqft;
    flags when a record reported sqft but all of it was implausible."""
    lo, hi = SQFT_BANDS.get(bedrooms, _SQFT_FALLBACK)
    reported = [u["sqft"] for u in units if isinstance(u.get("sqft"), (int, float)) and u["sqft"] > 0]
    good = [s for s in reported if lo <= s <= hi]
    if good:
        return _avg(good), ""
    if reported:                                   # had sqft, all out of band
        return None, f"reported {min(reported)}-{max(reported)} sf, implausible for {bedrooms}BR — dropped"
    return None, ""


def rollup_buildings(listings, subj_lat, subj_lon):
    """unit-level listings -> one comp record per building, with the Tier-A math."""
    groups = defaultdict(list)
    for r in listings:
        groups[_building_key(r)].append(r)

    comps = []
    for (addr, _zip), units in groups.items():
        lat = _avg([u["lat"] for u in units])
        lon = _avg([u["lon"] for u in units])
        rent = _avg([u["rent"] for u in units])
        beds_b = units[0].get("bedrooms")
        sqft, sqft_flag = _clean_sqft(units, beds_b)
        dist_mi = (ag.haversine_m(subj_lon, subj_lat, lon, lat) / M_PER_MILE
                   if lat and lon else None)
        corr = _max([u.get("correlation") for u in units])
        trend = next((u.get("rent_trend") for u in units if u.get("rent_trend")), "")
        comps.append({
            "address": units[0].get("formatted_address") or units[0]["address"],
            "city": units[0].get("city"),
            "lat": lat, "lon": lon,
            "distance_mi": round(dist_mi, 2) if dist_mi is not None else None,
            "n_listings": len(units),           # ESTIMATE basis for # units (JUDGMENT)
            "unit_size_sf": round(sqft) if sqft else None,
            "base_rent": round(rent) if rent else None,
            "value_ratio": round(rent / sqft, 2) if rent and sqft else None,
            "bedrooms": beds_b,
            "bathrooms": units[0].get("bathrooms"),
            "year_built": _max([u["year_built"] for u in units]),
            "rent_trend": trend,                # concession/trend signal from price history
            "correlation": round(corr, 3) if corr is not None else None,  # RentCast similarity (AVM)
            "sqft_flag": sqft_flag,             # set when reported sqft was implausible & dropped
            "listing_url": units[0].get("listing_url"),   # Zillow property page for detail pass
            "amenities_raw": units[0].get("amenities_raw") or [],
        })
    # AVM-curated comps (with a correlation score) rank first; ties broken by distance.
    comps.sort(key=lambda c: (-(c["correlation"] or 0), c["distance_mi"] is None,
                              c["distance_mi"] or 9e9))
    return comps


def _max(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return max(xs) if xs else None


def collect_comps(subject_addr, beds, radius_mi, top_n, demo=False, use_avm=False,
                  subj_sqft=None, source="ai"):
    """-> (subject_geo, {bed: {comps, estimate}}). Shortlists top_n per bed.

    source='rentcast' (default): RentCast API; source='ai': Firecrawl+Claude scraper.
    When source='ai' and --avm is requested the AVM flag is silently ignored
    (RentCast AVM is not available for the ai source)."""
    geo = {"matched_address": subject_addr, "lat": 34.2502, "lon": -118.5320} if demo \
        else geocode(subject_addr)
    subj_sqft = subj_sqft or {}
    zipcode = _extract_zip(geo.get("matched_address", ""))
    by_bed = {}
    for b in beds:
        estimate = None
        if demo:
            pool = (ai_scraper.demo_rentals(b) if source == "ai"
                    else rentcast.demo_rentals(b))
        elif source == "ai":
            pool = ai_scraper.search_rentals(geo["lat"], geo["lon"], b,
                                             zipcode=zipcode)
        else:
            pool = rentcast.search_rentals(geo["lat"], geo["lon"], b, radius_mi)
            if use_avm:
                avm = rentcast.avm_comps(geo["lat"], geo["lon"], b, sqft=subj_sqft.get(b))
                estimate = {"estimate": avm["estimate"], "range_low": avm["range_low"],
                            "range_high": avm["range_high"]}
                have = {p["id"] for p in pool}
                pool += [c for c in avm["comps"] if c["id"] not in have]
        comps = rollup_buildings(pool, geo["lat"], geo["lon"])
        shortlisted = comps[:top_n]
        if source == "ai" and not demo:
            ai_scraper.enrich_comps(shortlisted, b)
            # recompute value_ratio now that sqft is filled
            for c in shortlisted:
                if c.get("unit_size_sf") is None and c.get("sqft"):
                    c["unit_size_sf"] = c["sqft"]
                if c.get("base_rent") and c.get("unit_size_sf"):
                    c["value_ratio"] = round(c["base_rent"] / c["unit_size_sf"], 2)
        by_bed[b] = {"comps": shortlisted, "estimate": estimate}
    return geo, by_bed


# ---- grid layout (faithful to template/comp_grid_reference.xlsx) -------------
# Each tuple: (row label, fill key | None). None = analyst/judgment row left blank.
GRID_ROWS = [
    ("Distance in Miles from Subject", "distance_mi"),
    ("Number of Units", "n_listings"),          # JUDGMENT: listing-count estimate
    ("Vacancy Rate", None),
    ('Waiting List ("Y"es or "N"o)', None),
    ("Turnover Rate", None),
    ("Unit Size in SF", "unit_size_sf"),
    ("Base Rent", "base_rent"),
    ("Value Ratio ($/SF)", "value_ratio"),
    ("# of stories", None),                      # not in RentCast
    ('Elevator ("Y"es or "N"o)', None),
    ("# of Bedrooms", "bedrooms"),
    ("# of Bathrooms", "bathrooms"),
    ("Age (built or last renovated)", "year_built"),
    ("Rent Concessions", "rent_trend"),          # sourced from price history
    ("Comp similarity (RentCast 0-1)", "correlation"),   # diagnostic; AVM only
]
JUDGMENT_KEYS = {"n_listings"}


def write_grid(geo, by_bed, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr = Font(bold=True)
    subj_fill = PatternFill("solid", fgColor="DDEBF7")
    judg_fill = PatternFill("solid", fgColor="FFF2CC")

    for b, payload in by_bed.items():
        comps = payload["comps"]
        est = payload.get("estimate")
        ws = wb.create_sheet(BEDS_TO_SHEET.get(b, f"{b} BEDROOM MARKET"))
        ws.column_dimensions["A"].width = 34
        ws["A1"] = "Subject / Comp"; ws["A1"].font = hdr
        ws["B1"] = geo["matched_address"]; ws["B1"].font = hdr; ws["B1"].fill = subj_fill
        for i, c in enumerate(comps):
            col = 3 + i
            ws.cell(1, col, c["address"]).font = hdr
        ws["A2"] = "City"
        for i, c in enumerate(comps):
            ws.cell(2, 3 + i, c.get("city"))
        if est and est.get("estimate"):
            ws.cell(2, 2, f"RentCast est. ${est['estimate']:,} "
                          f"(${est['range_low']:,}-${est['range_high']:,})")

        r = 4
        for label, key in GRID_ROWS:
            ws.cell(r, 1, label).font = hdr
            for i, c in enumerate(comps):
                cell = ws.cell(r, 3 + i, c.get(key) if key else None)
                if key in JUDGMENT_KEYS and cell.value is not None:
                    cell.fill = judg_fill
            r += 1

        # weighted-average column (by unit-size SF), Tier-A computed
        avg_col = 3 + len(comps) + 1
        ws.cell(3, avg_col, "Wtd Avg").font = hdr
        wkeys = {"unit_size_sf", "base_rent", "value_ratio"}
        r = 4
        for label, key in GRID_ROWS:
            if key in wkeys:
                vals = [c.get(key) for c in comps if isinstance(c.get(key), (int, float))]
                if vals:
                    ws.cell(r, avg_col, round(sum(vals) / len(vals), 2))
            r += 1

        note_txt = (
            "Tier A (data) filled above. Yellow = JUDGMENT (unit count is a "
            "listing-count estimate). Adjustment $ columns, vacancy, turnover, "
            "utilities & amenities are analyst/site-visit fields — blank by design.")
        flagged = [c for c in comps if c.get("sqft_flag")]
        if flagged:
            note_txt += "  SQFT DROPPED (size/$ per SF left blank): " + "; ".join(
                f"{c['address']} ({c['sqft_flag']})" for c in flagged)
        note = ws.cell(r + 1, 1, note_txt)
        note.alignment = Alignment(wrap_text=True)
        note.font = Font(italic=True, size=9)

    wb.save(out_path)
    return out_path


# ---- formatted CTCAC grid (N1) — subject + comps with adjustments ------------
# Mirrors the SoLa "Stick Rent Comp" layout: column A labels, subject in B, each
# comp in a Char column with its Adj column to the right, then Adjusted Rent /
# value ratio / differential and the 110%-guardrail note. Adjustment $ come from
# comp_adjust (engine); blank subject characteristics leave those lines at 0 for
# the analyst (or the web editor) to fill.
def write_ctcac_grid(geo, by_bed, out_path, subjects=None, comp_chars=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import comp_adjust as CA

    subjects = subjects or {}
    comp_chars = comp_chars or {}
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    hdr = Font(bold=True); subj_fill = PatternFill("solid", fgColor="DDEBF7")
    adj_fill = PatternFill("solid", fgColor="FFF2CC"); ital = Font(italic=True, size=9)

    for b, payload in by_bed.items():
        comps = payload["comps"]
        subj = subjects.get(b, {})
        # map rollup records -> engine comp dicts. Amenity/utility chars come from
        # the editor (comp_chars) when present; RentCast itself carries none.
        cc = comp_chars.get(b, [])
        ecomps = [{"sf": c.get("unit_size_sf"), "rent": c.get("base_rent"),
                   "year": c.get("year_built"), "baths": c.get("bathrooms"),
                   "amenities": (cc[i]["amenities"] if i < len(cc) else {}),
                   "utilities": (cc[i]["utilities"] if i < len(cc) else {})}
                  for i, c in enumerate(comps)]
        esubj = {"sf": subj.get("sf"), "rent": subj.get("rent"), "year": subj.get("year"),
                 "baths": subj.get("baths"), "amenities": subj.get("amenities", {}),
                 "utilities": subj.get("utilities", {})}
        assessed = CA.assess(esubj, ecomps)

        ws = wb.create_sheet(BEDS_TO_SHEET.get(b, f"{b} BEDROOM MARKET"))
        ws.column_dimensions["A"].width = 30
        ws["A1"] = "Enter Data in Blue or Blank Fields"; ws["A1"].font = hdr
        ws["B1"] = geo["matched_address"]; ws["B1"].font = hdr; ws["B1"].fill = subj_fill
        for i, c in enumerate(comps):
            ws.cell(1, 3 + 2 * i, c["address"]).font = hdr   # char column header
        # static rows: (label, subject value, comp-key | None)
        rows = [
            ("City", subj.get("city"), "city"),
            ('("M"arket, "L"ow Income)', subj.get("m_or_l", "M"), None),
            ("Distance in Miles from Subject", None, "distance_mi"),
            ("Number of Units", subj.get("units"), "n_listings"),
            ("Unit Size in SF", subj.get("sf"), "unit_size_sf"),
            ("Base Rent", subj.get("rent"), "base_rent"),
            ("Value Ratio ($/SF)", None, "value_ratio"),
            ("# of Bedrooms", b, "bedrooms"),
            ("# of Bathrooms", subj.get("baths"), "bathrooms"),
            ("Age (built or last renovated)", subj.get("year"), "year_built"),
        ]
        r = 3
        for label, sval, key in rows:
            ws.cell(r, 1, label).font = hdr
            if sval is not None:
                ws.cell(r, 2, sval)
            for i, c in enumerate(comps):
                ws.cell(r, 3 + 2 * i, c.get(key) if key else None)
            r += 1

        # adjustment lines (Char in comp col, Adj $ in the next col) from the engine
        r += 1
        ws.cell(r, 1, "ADJUSTMENTS ($)").font = hdr; r += 1
        adj_labels = [lab for lab, _ in assessed[0]["adjustments"]] if assessed else []
        for li, label in enumerate(adj_labels):
            ws.cell(r, 1, label)
            for i, a in enumerate(assessed):
                val = a["adjustments"][li][1]
                cell = ws.cell(r, 4 + 2 * i, val)        # adj column
                if val:
                    cell.fill = adj_fill
            r += 1

        # results
        r += 1
        for label, field, getter in [
            ("Adjusted Rent", "adjusted_rent", lambda a: a["adjusted_rent"]),
            ("Adjusted Value Ratio ($/SF)", "adjusted_ratio", lambda a: a["adjusted_ratio"]),
            ("*Adjusted Rent / Base Rent", "ratio_to_base", lambda a: a["ratio_to_base"]),
        ]:
            ws.cell(r, 1, label).font = hdr
            for i, a in enumerate(assessed):
                c = ws.cell(r, 4 + 2 * i, getter(a))
                if field == "ratio_to_base" and a["over_guardrail"]:
                    c.fill = PatternFill("solid", fgColor="FCE4D6")
            r += 1
        note = ws.cell(r + 1, 1,
                       "*Comps exceeding 110% of adjusted rent/base rent must be justified. "
                       "Adjustment $ are engine defaults (size = subj $/SF x 0.10; age = $5/yr; "
                       "per-line amenity/utility $) — edit per appraiser judgment.")
        note.alignment = Alignment(wrap_text=True); note.font = ital
    wb.save(out_path)
    return out_path


def _print_summary(geo, by_bed):
    print(f"\nSubject: {geo['matched_address']}  ({geo['lat']:.4f}, {geo['lon']:.4f})")
    for b, payload in by_bed.items():
        comps, est = payload["comps"], payload.get("estimate")
        head = f"\n=== {BEDS_TO_SHEET.get(b, str(b)+'BR')}  —  {len(comps)} comp(s)"
        if est and est.get("estimate"):
            head += f"  |  RentCast est. ${est['estimate']:,} (${est['range_low']:,}-${est['range_high']:,})"
        print(head + " ===")
        if not comps:
            print("  (no listings found in radius)")
            continue
        print(f"  {'address':38} {'dist':>5} {'#u':>3} {'sqft':>5} {'rent':>6} {'$/sf':>5} {'yr':>5} {'corr':>5}  trend")
        for c in comps:
            flag = " !sqft" if c.get("sqft_flag") else ""
            print(f"  {(c['address'] or '')[:38]:38} {c['distance_mi'] or 0:5.2f} "
                  f"{c['n_listings']:>3} {c['unit_size_sf'] or 0:>5} {c['base_rent'] or 0:>6} "
                  f"{c['value_ratio'] or 0:>5} {c['year_built'] or 0:>5} "
                  f"{c['correlation'] or 0:>5}  {c['rent_trend'] or ''}{flag}")
        flagged = [c for c in comps if c.get("sqft_flag")]
        for c in flagged:
            print(f"  ! sqft dropped: {c['address']} — {c['sqft_flag']}")


def main():
    ap = argparse.ArgumentParser(description="Collect rent comps into a CTCAC-style grid.")
    ap.add_argument("subject", help="subject property address")
    ap.add_argument("--beds", type=int, nargs="+", default=[1], help="bed counts to comp")
    ap.add_argument("--radius", type=float, default=rentcast.DEFAULT_RADIUS_MI, help="search radius (mi)")
    ap.add_argument("--top", type=int, default=3, help="comps to shortlist per bed count")
    ap.add_argument("--avm", action="store_true",
                    help="also pull RentCast rent-AVM comps (adds similarity score + rent estimate; rentcast source only)")
    ap.add_argument("--source", choices=["rentcast", "ai"], default="ai",
                    help="comp data source: 'ai' (Firecrawl+Gemini, default) or 'rentcast' (API)")
    ap.add_argument("--demo", action="store_true", help="use offline fixture (no API keys needed)")
    ap.add_argument("--ctcac", action="store_true",
                    help="write the formatted CTCAC grid (subject + comps + adjustments) instead of the clean grid")
    ap.add_argument("--out", default="build/_comps_demo.xlsx", help="output xlsx path")
    args = ap.parse_args()

    geo, by_bed = collect_comps(args.subject, args.beds, args.radius, args.top,
                                demo=args.demo, use_avm=args.avm, source=args.source)
    _print_summary(geo, by_bed)
    path = (write_ctcac_grid(geo, by_bed, args.out) if args.ctcac
            else write_grid(geo, by_bed, args.out))
    print(f"\nGrid written -> {path}")


if __name__ == "__main__":
    main()
