#!/usr/bin/env python3
"""
collect.py — run every Tier-A reader against a property workbook through
runner.run_field. This is the keystone (geocoder) feeding the Tier-A source
modules into the state/log loop: each field gets its answer (col C), notes
(col E), a STATE (col A), and a State Log row, with TOOL-FAIL -> MANUAL-VERIFY
escalation handled by run_field.

Usage:
  python build/collect.py --demo "11300 S Main St, Los Angeles, CA 90061"
      copy template/ to a throwaway workbook, geocode the address, run all
      readers, and print the resulting states + cells.
  python build/collect.py path/to/property.xlsx [address]
      run against a real per-property workbook (address read from C3 if omitted).
  python build/collect.py path/to/property.xlsx "11070 Borden Ave; 11080 Borden Ave; ..."
      ASSEMBLAGE — run several addresses as ONE site (separate with ';'). Point /
      tract readers run on the primary (first) parcel; land_sf is SUMMED and every
      APN listed across all parcels. Flags if the parcels span multiple census
      tracts (which makes the tract-based fields — QCT/DDA/resource — ambiguous).
      This is the address-based alternative to identifying a multi-parcel site.

A field is automated only if it appears in READERS; everything else keeps the
state the template/generator assigned (DESK-PENDING, BROWSER-PENDING, etc.).
"""
import sys, json, shutil, datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build" / "sources"))

import yaml
from openpyxl import load_workbook
from runner import run_reader, apply_outcome
from geocoder import geocode
import fema, hud, tcac, oz, calfire, calgem, cgs, ust, zimas, jurisdiction, parcel, nc, pha
import places, slope, towers, airport, coastal, sandiego
from jurisdiction import _county_basename

# field id -> reader callable(geo) -> {"answer","notes"} (or raises -> TOOL-FAIL)
# Statewide / federal / derived sources — run for any CA parcel.
READERS = {
    "address": jurisdiction.address,
    "apn": jurisdiction.apn,
    "city_jurisdiction": jurisdiction.city_jurisdiction,
    "county": jurisdiction.county,
    "geographic_pool": jurisdiction.geographic_pool,
    "pha": pha.pha,
    "land_sf": parcel.land_sf,
    "qct": hud.qct,
    "dda": hud.dda,
    "resource_area": tcac.resource_area,
    "neighborhood_change_area": nc.neighborhood_change,
    "opportunity_zone": oz.opportunity_zone,
    "flood_zone": lambda geo: fema.flood_zone(geo["lon"], geo["lat"]),
    "very_high_fire_hazard_zone": calfire.very_high_fire_hazard_zone,
    "airport_hazard_zone": airport.airport_hazard_zone,
    "coastal_zone": coastal.coastal_zone,
    "wells_on_site": calgem.wells_on_site,
    "liquefaction_zone": cgs.liquefaction_zone,
    "alquist_priolo_fault_zone": cgs.alquist_priolo_fault_zone,
    "underground_storage_tanks": ust.underground_storage_tanks,
    "slope_grade": slope.slope_grade,
    "cell_towers": towers.cell_towers,
    # proximity (OpenStreetMap)
    "nearest_bus_stop": places.nearest_bus_stop,
    "nearest_grocery_store": places.nearest_grocery_store,
    "nearest_park": places.nearest_park,
    "nearest_medical_clinic": places.nearest_medical_clinic,
    "nearest_library": places.nearest_library,
    "nearest_pharmacy": places.nearest_pharmacy,
    "nearest_school": places.nearest_school,
    "nearest_qualifying_transit_stop": places.nearest_qualifying_transit_stop,
}

# LA-City zoning/hazard block (ZIMAS-equivalent) — only run when the parcel is
# in LA City; off LA City these route to local planning and stay manual.
ZIMAS_READERS = {
    "zoning": zimas.zoning,
    "q_conditions_la": zimas.q_conditions,
    "specific_plan_overlay": zimas.specific_plan_overlay,
    "council_supervisor_district": zimas.council_district,
    "historic_status": zimas.historic_status,
    "methane_hazard_zone_la": zimas.methane_hazard_zone,
    "toc_tier_la": zimas.toc_tier,
    "half_mile_major_transit": zimas.half_mile_major_transit,
    "transitional_height_adj_zones": zimas.transitional_height,   # derived -> JUDGMENT
    "tier_transit_verification": zimas.tier_transit_verification,
}

# City-of-San-Diego municipal block (SD analog of ZIMAS) — only run when the
# parcel is in San Diego County; off the City of SD these readers raise and the
# fields route to local planning. (airport_hazard_zone is NOT here — it stays in
# the always-run READERS, with an SD branch inside airport.py.)
# q_conditions_la / methane_hazard_zone_la / transitional_height_adj_zones are
# LA-only concepts -> N/A in SD (no readers).
SD_READERS = {
    "zoning": sandiego.zoning,
    "specific_plan_overlay": sandiego.specific_plan_overlay,
    "council_supervisor_district": sandiego.council_district,
    "historic_status": sandiego.historic_status,
    "toc_tier_la": sandiego.transit_priority_area,   # SD Transit Priority Area in the TOC/Tier cell
    "half_mile_major_transit": sandiego.half_mile_major_transit,
    "tier_transit_verification": sandiego.tier_transit_verification,
}


def _parse_addresses(s):
    """A ';'-separated string -> assemblage list; a single address -> itself."""
    if isinstance(s, str) and ";" in s:
        return [a.strip() for a in s.split(";") if a.strip()]
    return s


def _assemble_parcels(geos):
    """Aggregate the parcel-level fields across an assemblage of addresses run as
    one site: SUM gross lot area over the unique parcels and list every APN.
    Returns {field_id: outcome} (collect applies them, overriding the primary's).
    Point/tract fields are NOT aggregated — they run on the primary parcel."""
    parts, fails, seen = [], [], set()
    for g in geos:
        addr = g.get("matched_address")
        try:
            ap = jurisdiction.apn(g)["answer"]
        except Exception:
            ap = None
        try:
            lo = parcel.land_sf(g)
        except Exception as e:
            fails.append((ap or addr, str(e)))
            continue
        key = ap or addr
        if key in seen:                 # two addresses on the same parcel -> count once
            continue
        seen.add(key)
        parts.append({"addr": addr, "apn": ap, "area": lo["answer"],
                      "matched": lo.get("state", "VERIFIED") == "VERIFIED"})

    clean = bool(parts) and all(p["matched"] for p in parts) and not fails
    state = "VERIFIED" if clean else "JUDGMENT"
    out = {}
    out["address"] = ("ok", {"answer": "; ".join(p["addr"] for p in parts)
                             or "; ".join(g.get("matched_address", "?") for g in geos),
                             "notes": f"Assemblage of {len(geos)} addresses run as one site. "
                                      f"Source: U.S. Census geocoder."})
    apns = [p["apn"] for p in parts if p["apn"]]
    out["apn"] = ("ok", {"answer": "; ".join(apns) if apns else "(unresolved)", "state": state,
                         "notes": f"{len(apns)} parcels in assemblage: {', '.join(apns) or 'none resolved'}. "
                                  f"Source: Assessor parcels."})
    if not parts:
        out["land_sf"] = ("fail", LookupError("no assemblage parcels could be sized"))
    else:
        total = sum(p["area"] for p in parts)
        lines = "; ".join(f"{p['apn'] or p['addr']}={p['area']:,} sf" for p in parts)
        note = (f"Assemblage gross land area {total:,} sf = sum of {len(parts)} parcels [{lines}]. "
                f"GROSS — reconcile buildable vs ALTA survey. Source: LA City/County Assessor parcels.")
        if fails:
            note += " UNSIZED: " + "; ".join(a for a, _ in fails) + " — verify."
        out["land_sf"] = ("ok", {"answer": total, "state": state, "notes": note})
    return out


def collect(wb_path, address, property_id=None, workers=10):
    schema = yaml.safe_load((ROOT / "canonical/schema.yaml").read_text())
    by_id = {f["id"]: f for f in schema["fields"]}
    if isinstance(address, str):
        address = _parse_addresses(address)        # "a; b" -> ["a", "b"] (assemblage); single -> str
    addresses = list(address) if isinstance(address, (list, tuple)) else [address]
    geos = [geocode(a) for a in addresses]
    geo = geos[0]                       # primary parcel drives all point/tract readers
    multi = len(geos) > 1
    print(f"geocoded: {geo['matched_address']}  tract {geo['geoid']}  "
          f"({geo['lat']:.5f},{geo['lon']:.5f})")
    if multi:
        print(f"ASSEMBLAGE: {len(geos)} addresses run as one site (primary above):")
        for g in geos[1:]:
            print(f"  + {g['matched_address']}  tract {g['geoid']}")
    print()

    active = dict(READERS)
    if multi:                          # parcel fields are aggregated across the assemblage, not snapped to primary
        for fid in ("address", "apn", "land_sf"):
            active.pop(fid, None)
    if zimas.in_la_city(geo):       # also warms the parcel snap the ZIMAS readers share
        active.update(ZIMAS_READERS)
        print("primary parcel is in LA City -> running ZIMAS block\n")
    elif _county_basename(geo) == "San Diego":
        active.update(SD_READERS)
        print("primary parcel is in San Diego -> running SD block\n")
    else:
        print("primary parcel not in LA City / San Diego -> municipal block skipped (route to local planning)\n")
    for _warm in (nc._load, tcac._load_index):   # warm the shared bulk caches once (thread-safe)
        try:
            _warm()
        except Exception:
            pass

    # Phase 1 — fetch every reader concurrently (I/O-bound; off the workbook).
    def _fetch(item):
        fid, fn = item
        return fid, run_reader(lambda: fn(geo))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        outcomes = dict(ex.map(_fetch, list(active.items())))

    # Assemblage — aggregate the parcel fields across all addresses + flag tract divergence.
    if multi:
        agg = _assemble_parcels(geos)
        tracts = sorted({g["geoid"] for g in geos})
        if len(tracts) > 1:
            warn = (f"ASSEMBLAGE SPANS {len(tracts)} CENSUS TRACTS ({', '.join(tracts)}) — "
                    f"tract-based fields (QCT/DDA/resource/opportunity zone/neighborhood change) "
                    f"reflect the PRIMARY parcel only; verify per parcel.")
            print("  ⚠ " + warn + "\n")
            agg["address"][1]["notes"] += " " + warn
        outcomes.update(agg)

    # Phase 2 — apply all outcomes to the workbook in a single open/save.
    wb = load_workbook(wb_path)
    ws, log = wb["Site DD"], wb["State Log"]
    ts = datetime.datetime.now().isoformat(timespec="seconds")
    results = {}
    for fid in outcomes:
        field = by_id[fid]
        state = apply_outcome(ws, log, field, outcomes[fid], property_id=property_id or "DEMO", ts=ts)
        results[fid] = (field["answer_cell"], state)
        print(f"  {field['answer_cell']:5} {fid:28} -> {state}")
    wb.save(wb_path)
    return results, geo


def _dump(wb_path, results):
    from openpyxl import load_workbook
    ws = load_workbook(wb_path)["Site DD"]
    print("\n--- written cells (Status A | answer C | notes E) ---")
    for fid, (cell, _) in results.items():
        row = int(cell[1:])
        print(f"  {fid:28} A={ws.cell(row,1).value!s:14} C={ws.cell(row,3).value!s:18} "
              f"E={(ws.cell(row,5).value or '')[:80]}")


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--demo":
        address = _parse_addresses(" ".join(args[1:]) or "11300 S Main St, Los Angeles, CA 90061")
        tmp = ROOT / "build" / "_collect_demo.xlsx"
        shutil.copy(ROOT / "template/Checklist_BLANK_master.xlsx", tmp)
        results, _ = collect(tmp, address)
        _dump(tmp, results)
        print(f"\nworkbook: {tmp}")
    elif args:
        from openpyxl import load_workbook
        wb = args[0]
        address = _parse_addresses(" ".join(args[1:]) or load_workbook(wb)["Site DD"]["C3"].value)
        results, _ = collect(wb, address)
        _dump(wb, results)
    else:
        print(__doc__)
