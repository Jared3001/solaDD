#!/usr/bin/env python3
"""
runner.py — the read -> state -> log loop. This is the state taxonomy made
executable, and the reference pattern every automated source plugs into.

Given a property workbook, a schema field, and a reader() callable:
  - reader succeeds -> write answer to Column C, append notes to Column E,
    set Status (Column A) = VERIFIED, log success.
  - reader raises -> increment the TOOL-FAIL counter; at max_attempts it
    auto-escalates to MANUAL-VERIFY. Every attempt appends a State Log row.

reader() returns {"answer": <cell value>, "notes": <str>} or raises on failure.
No network needed here — the escalation control flow is fully testable offline
with a mock reader (see --selftest).
"""
import datetime
from openpyxl import load_workbook

MAX_ATTEMPTS = 3   # mirrors taxonomy.yaml escalation.max_attempts

# Close/landing states a reader may request via {"state": ...} on success.
# Default is VERIFIED (primary-source confirmed). JUDGMENT lets a reader that
# can only DERIVE a borderline answer (e.g. transitional height) route it up for
# senior review instead of overclaiming VERIFIED. NA for "rule does not apply".
READER_LANDING_STATES = {"VERIFIED", "JUDGMENT", "NA", "COMPUTED", "OM-SOURCED"}


def _attempt_count(token) -> int:
    if isinstance(token, str) and token.startswith("TOOL-FAIL"):
        try:
            return int(token.split()[1].split("/")[0])
        except Exception:
            return 0
    return 0


def run_field(wb_path, field: dict, reader, *, property_id=None, max_attempts=MAX_ATTEMPTS) -> str:
    wb = load_workbook(wb_path)
    ws, log = wb["Site DD"], wb["State Log"]
    row = int(field["answer_cell"][1:])
    status = ws.cell(row, 1)
    n = _attempt_count(status.value)
    ts = datetime.datetime.now().isoformat(timespec="seconds")
    pid = property_id or ws["C3"].value or str(wb_path)
    tool = field.get("source_of_record", "")
    try:
        out = reader()
        ws.cell(row, 3, out["answer"])
        if out.get("notes"):
            prev = ws.cell(row, 5).value
            ws.cell(row, 5, (prev + " | " if prev else "") + out["notes"])
        landing = out.get("state", "VERIFIED")
        if landing not in READER_LANDING_STATES:
            landing = "VERIFIED"
        status.value = landing
        log.append([ts, pid, field["id"], tool, "success", n + 1, ""])
    except Exception as e:
        n += 1
        if n >= max_attempts:
            status.value = "MANUAL-VERIFY"
            log.append([ts, pid, field["id"], tool, "fail", n, f"escalated after {n}: {e}"[:200]])
        else:
            status.value = f"TOOL-FAIL {n}/{max_attempts}"
            log.append([ts, pid, field["id"], tool, "fail", n, str(e)[:200]])
    wb.save(wb_path)
    return status.value


def _selftest():
    """Offline proof of the escalation control flow — no network."""
    import shutil, yaml
    from pathlib import Path
    root = Path(__file__).resolve().parent.parent
    schema = yaml.safe_load((root / "canonical/schema.yaml").read_text())
    flood = next(f for f in schema["fields"] if f["id"] == "flood_zone")
    tmp = root / "build/_selftest.xlsx"
    shutil.copy(root / "template/Checklist_BLANK_master.xlsx", tmp)

    def failing():
        raise RuntimeError("viewer unresponsive")

    def succeeding():
        return {"answer": "No", "notes": "Zone X — outside SFHA. Source: FEMA NFHL REST."}

    seq = [run_field(tmp, flood, failing, property_id="TEST") for _ in range(3)]
    seq.append(run_field(tmp, flood, succeeding, property_id="TEST"))
    print("escalation sequence:", seq)
    assert seq == ["TOOL-FAIL 1/3", "TOOL-FAIL 2/3", "MANUAL-VERIFY", "VERIFIED"], seq

    def deriving():   # a reader that can only derive a borderline answer
        return {"answer": "Likely applies", "notes": "derived", "state": "JUDGMENT"}

    judged = run_field(tmp, flood, deriving, property_id="TEST")
    print("reader-requested landing state:", judged)
    assert judged == "JUDGMENT", judged

    from openpyxl import load_workbook
    log = load_workbook(tmp)["State Log"]
    rows = [[c.value for c in r] for r in log.iter_rows(min_row=2) if r[0].value]
    print(f"State Log rows written: {len(rows)}")
    for r in rows:
        print("  ", r[2], r[4], f"attempt {r[5]}", "-", r[6] or "ok")
    tmp.unlink()
    print("SELFTEST PASS — TOOL-FAIL counts, escalates at 3, then resolves to VERIFIED.")


if __name__ == "__main__":
    import sys
    if "--selftest" in sys.argv:
        _selftest()
    else:
        print("usage: runner.py --selftest   (live field runs are driven by the per-source modules)")
