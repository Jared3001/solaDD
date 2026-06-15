#!/usr/bin/env python3
"""
validate.py — enforce that the blank master agrees with the canonical schema.

This is the mechanism that makes "schema is the single source of truth" real:
if anyone hand-edits the master out of sync with canonical/schema.yaml, this
fails. Run it in CI on every change and before cutting a new template.

Usage:
    python build/validate.py [path/to/master.xlsx]   (defaults to template/)
Exit code 0 = in sync, 1 = drift found.
"""
import sys, yaml
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple

ROOT = Path(__file__).resolve().parent.parent
SCHEMA = yaml.safe_load((ROOT / "canonical/schema.yaml").read_text())
TAX    = yaml.safe_load((ROOT / "canonical/taxonomy.yaml").read_text())
MASTER = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "template/Checklist_BLANK_master.xlsx"

wb = load_workbook(MASTER, data_only=False)
ws = wb["Site DD"]
errors, checks = [], 0

# dropdown map: (row,col) -> formula1 string
dv_map = {}
for dv in ws.data_validations.dataValidation:
    for rng in str(dv.sqref).split():
        if ":" in rng:
            c1, r1, c2, r2 = range_boundaries(rng)
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    dv_map[(r, c)] = dv.formula1
        else:
            dv_map[coordinate_to_tuple(rng)] = dv.formula1

valid_states = {s["id"] for s in TAX["states"]}
dsm = TAX["default_state_map"]

for f in SCHEMA["fields"]:
    r = f["row"]

    # 1) label matches
    checks += 1
    got = (ws.cell(r, 2).value or "").strip()
    if got != f["label"].strip():
        errors.append(f"row {r}: label '{got}' != schema '{f['label']}'")

    # 2) dropdown matches (enum / yes_no)
    if "dropdown_verbatim" in f:
        checks += 1
        exp = '"' + f["dropdown_verbatim"] + '"'
        if dv_map.get((r, 3)) != exp:
            errors.append(f"row {r} ({f['id']}): dropdown {dv_map.get((r,3))!r} != schema {exp!r}")
    elif f.get("answer_type") == "yes_no" and f.get("allowed_values_status") == "confirmed":
        checks += 1
        if dv_map.get((r, 3)) != '"Yes,No"':
            errors.append(f"row {r} ({f['id']}): expected Yes/No dropdown, got {dv_map.get((r,3))!r}")

    # 3) formulas match
    if "formula" in f:
        checks += 1
        if ws.cell(r, 3).value != f["formula"]:
            errors.append(f"row {r} ({f['id']}): formula {ws.cell(r,3).value!r} != schema {f['formula']!r}")

    # 4) blank template: answer cell empty unless computed
    if f.get("answer_type") != "computed":
        checks += 1
        v = ws.cell(r, 3).value
        if v not in (None, ""):
            errors.append(f"row {r} ({f['id']}): blank template has leftover answer {v!r} in C")

    # 5) status gutter (col A) initialized to the schema's default state
    checks += 1
    a = (ws.cell(r, 1).value or "").strip()
    want = dsm.get(f["default_state"])
    if a not in valid_states:
        errors.append(f"row {r} ({f['id']}): status '{a}' not a taxonomy state")
    elif a != want:
        errors.append(f"row {r} ({f['id']}): status '{a}' != expected '{want}'")

print(f"schema v{SCHEMA['meta']['schema_version']} | taxonomy v{TAX['meta']['taxonomy_version']}")
print(f"checked {len(SCHEMA['fields'])} fields, {checks} assertions")
if errors:
    print(f"\nDRIFT — {len(errors)} problem(s):")
    for e in errors[:50]:
        print("  -", e)
    sys.exit(1)
print("PASS — master is in sync with canonical schema.")
