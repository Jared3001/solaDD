#!/usr/bin/env python3
"""
generate.py — build a blank master .xlsx from the canonical schema + taxonomy.

This realizes "the master is generated from the schema": rows, labels, section
order, dropdown menus, formulas, the Status gutter (initialized from the
taxonomy's default_state_map), the CDLAC lookup, and a per-property State Log
tab all come from canonical/. 

Scope note: this generates the STRUCTURAL + DATA layer (the drift-prone parts).
It does NOT reproduce cosmetic scaffolding the schema deliberately doesn't own —
the '[Link ...]' placeholders (col D) and 'Guidelines/Resources' hint text (col F).
Those live in the styled skeleton at template/ . Workflow: edit schema -> run
validate.py against the styled template; only regenerate from scratch when you
want a clean structural rebuild.

Usage:  python build/generate.py [out.xlsx]
"""
import sys, csv, yaml
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
SCHEMA = yaml.safe_load((ROOT / "canonical/schema.yaml").read_text())
TAX    = yaml.safe_load((ROOT / "canonical/taxonomy.yaml").read_text())
OUT    = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "template/Checklist_BLANK_master_generated.xlsx"

dsm = TAX["default_state_map"]
sections = {s["id"]: s for s in SCHEMA["sections"]}

wb = Workbook()
ws = wb.active
ws.title = "Site DD"
bold = Font(bold=True)

# version stamp + header identity block
ws["B2"] = "DD Checklist"; ws["B2"].font = bold
ws["E2"] = f"schema v{SCHEMA['meta']['schema_version']} · taxonomy v{TAX['meta']['taxonomy_version']}"
for r, lbl in [(3, "ADDRESS"), (4, "APN"), (5, "Acq Lead:"), (6, "Dev Lead:"), (7, "Date last updated:")]:
    ws.cell(r, 2, lbl)

# section header bands
band = {3: "Details/Condition", 4: "Supporting Documentation", 5: "Notes",
        6: "Guidelines/Resources", 7: "Acq", 8: "Date", 9: "Dev", 10: "Date"}
for s in SCHEMA["sections"]:
    r = s["start_row"]
    ws.cell(r, 1, "Status").font = bold
    ws.cell(r, 2, s["label"]).font = bold
    for col, txt in band.items():
        ws.cell(r, col, txt).font = bold

# fields: label, formula/blank answer, status gutter, dropdown
dv_objs = {}  # formula1 -> DataValidation
def dv_for(formula1):
    if formula1 not in dv_objs:
        d = DataValidation(type="list", formula1=formula1, allowBlank=True)
        ws.add_data_validation(d); dv_objs[formula1] = d
    return dv_objs[formula1]

for f in SCHEMA["fields"]:
    r = f["row"]
    ws.cell(r, 2, f["label"])
    if "formula" in f:
        ws.cell(r, 3, f["formula"])
    ws.cell(r, 1, dsm.get(f["default_state"], ""))   # status gutter (all fields)
    if "dropdown_verbatim" in f:
        dv_for('"' + f["dropdown_verbatim"] + '"').add(ws.cell(r, 3))
    elif f.get("answer_type") == "yes_no" and f.get("allowed_values_status") == "confirmed":
        dv_for('"Yes,No"').add(ws.cell(r, 3))

ws.column_dimensions["A"].width = 17
ws.column_dimensions["B"].width = 34

# CDLAC lookup sheet from canonical data
lk = wb.create_sheet("Sheet1")
lk["B2"] = "County"; lk["C2"] = "CDLAC Region"
with (ROOT / "canonical/cdlac_regions.csv").open() as fh:
    for i, row in enumerate(csv.DictReader(fh), start=3):
        lk.cell(i, 2, row["County"]); lk.cell(i, 3, row["CDLAC Region"])

# per-property State Log tab (decision: per-property, not central)
log = wb.create_sheet("State Log")
for j, c in enumerate(TAX["state_log"]["columns"], 1):
    log.cell(1, j, c).font = bold

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"generated {OUT}  (schema v{SCHEMA['meta']['schema_version']}, taxonomy v{TAX['meta']['taxonomy_version']})")
print("note: run scripts/recalc on it if your toolchain needs cached formula values.")
