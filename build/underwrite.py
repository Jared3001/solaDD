#!/usr/bin/env python3
"""
underwrite.py — first-pass underwriting exporter.

Turns a completed DD checklist into TWO ready-to-recalc pro-forma models
(Stick + Modular) by writing the DD-derived site inputs and the confirmed
projection-logic assumptions (UNDERWRITING_INTAKE.md Part B) into SoLa's
pro-forma template, preserving its macros / LAMBDA names / formulas.

The exporter only sets Auto (from DD) + Logic cells. It deliberately leaves the
analyst's Hand fields untouched: residential stories (Pro_Forma!C15, which the
C9 construction-type formula keys on), acquisition price, BIPOC, prevailing
wage. So the TEMPLATE should be a clean master pro-forma, not a filled deal —
otherwise those hand fields carry over.  --template defaults to $SOLA_UW_TEMPLATE.

Usage:
  python build/underwrite.py path/to/DD_checklist.xlsx \
      [--template path/to/master.xlsm] [--out DIR] [--name "Deal Name"]

  # round-trip / self-check against the analyzed example:
  python build/underwrite.py --selftest path/to/"Stick ... Kinzie Street.xlsm"
"""
import os, sys, argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build" / "sources"))

import yaml
from openpyxl import load_workbook
import uw_logic

DD_SHEET = "Site DD"
METHODS = ("Stick", "Modular")
# clean master pro-forma shipped in the repo (deal-specific hard inputs wiped:
# stories C15, acquisition S16, plus the exporter-overwritten id cells). Used when
# --template / $SOLA_UW_TEMPLATE aren't given.
DEFAULT_TEMPLATE = ROOT / "template" / "ProForma_BLANK_master.xlsm"


def _dd_cell_map():
    """field id -> DD-workbook answer cell, from canonical/schema.yaml."""
    schema = yaml.safe_load((ROOT / "canonical/schema.yaml").read_text())
    return {f["id"]: f.get("answer_cell") for f in schema["fields"]}


# the DD outputs the exporter consumes
DD_FIELDS = ["address", "land_sf", "county", "pha", "qct", "dda",
             "resource_area", "neighborhood_change_area", "city_jurisdiction"]


def read_dd(dd_path):
    """Read the DD checklist's answers (Site DD col C) into {field: value}."""
    cells = _dd_cell_map()
    ws = load_workbook(dd_path, data_only=True)[DD_SHEET]
    out = {}
    for fid in DD_FIELDS:
        ref = cells.get(fid)
        out[fid] = ws[ref].value if ref else None
    return out


def export(dd, template_path, out_dir, deal_name=None):
    """Write <deal> — Stick.xlsm and <deal> — Modular.xlsm. Returns [paths]."""
    base, meta = uw_logic.base_cells(dd)
    deal = deal_name or base[("Pro_Forma", "B2")] or "Untitled Deal"
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    written = []
    for method in METHODS:
        wb = load_workbook(template_path, keep_vba=True)   # preserves macros + LAMBDAs
        cells = {**base, **uw_logic.method_cells(method)}
        for (sheet, ref), val in cells.items():
            wb[sheet][ref] = val
        out = out_dir / f"{deal} — {method}.xlsm"        # em dash per Part C spec
        wb.save(out)
        written.append(out)
    return written, meta


# ---------------- self-test: round-trip the analyzed Kinzie example ----------------
def selftest(example_xlsm):
    """Read the example .xlsm's own inputs as a pseudo-DD, regenerate both models,
    and assert the written cells match what uw_logic computed."""
    v = load_workbook(example_xlsm, data_only=True)["Pro_Forma"]
    # reconstruct a DD-style dict from the example's filled model inputs
    dd = {
        "address": v["B2"].value,
        "county": v["C3"].value,
        "pha": v["C4"].value,
        "qct": "Yes" if v["C5"].value == "QCT" else ("Yes" if v["C5"].value == "DDA" else "No"),
        "dda": "Yes" if v["C5"].value == "DDA" else "No",
        "resource_area": {"Highest": "Highest Resource", "High": "High Resource",
                          "Medium": "Moderate Resource", "Low": "Low Resource"}.get(v["C6"].value, v["C6"].value),
        "neighborhood_change_area": v["C7"].value,
        "land_sf": v["C12"].value,
        "city_jurisdiction": v["C3"].value,
    }
    out_dir = ROOT / "build" / "_uw_selftest"
    paths, meta = export(dd, example_xlsm, out_dir, deal_name=uw_logic.project_name(dd["address"]))

    print(f"product: {meta['product']}  resource->{meta['resource_mapped']}")
    if meta["flags"]:
        for f in meta["flags"]:
            print(f"  FLAG: {f}")
    ok = True
    for p in paths:
        method = "Modular" if "Modular" in p.name else "Stick"
        wb = load_workbook(p, data_only=False)
        pf, dm = wb["Pro_Forma"], wb["Draws_Module"]
        checks = {
            "B2": uw_logic.project_name(dd["address"]),
            "C3": "Los Angeles", "C5": "QCT", "C6": "Highest", "C8": "No",
            "C9": uw_logic.C9_FORMULA,
            "I5": 0.25, "I6": 0.25, "R35": 0.10, "R36": 0.10, "R38": 0,
            "A36": method,
        }
        if method == "Modular":
            checks.update({"L5": 804, "L6": 994})
        for ref, exp in checks.items():
            got = pf[ref].value
            flag = "ok" if got == exp else "MISMATCH"
            if got != exp:
                ok = False
            print(f"  [{method}] {ref}: got={got!r} exp={exp!r} {flag}")
        bt = dm["B5"].value
        exp_bt = uw_logic.CONSTRUCTION_TIME[method]
        if bt != exp_bt:
            ok = False
        print(f"  [{method}] Draws_Module!B5 (build time): got={bt!r} exp={exp_bt} "
              f"{'ok' if bt == exp_bt else 'MISMATCH'}")
        # macros preserved?
        import zipfile
        with zipfile.ZipFile(p) as z:
            vba = any("vbaProject" in n for n in z.namelist())
        print(f"  [{method}] macros preserved: {vba}   -> {p}")
        if not vba:
            ok = False
    print("\nSELFTEST:", "PASS" if ok else "FAIL")
    return ok


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("dd", nargs="?", help="DD checklist workbook (.xlsx) OR, with --selftest, the example .xlsm")
    ap.add_argument("--template",
                    default=os.environ.get("SOLA_UW_TEMPLATE") or (str(DEFAULT_TEMPLATE) if DEFAULT_TEMPLATE.exists() else None),
                    help="master pro-forma .xlsm (default: $SOLA_UW_TEMPLATE, else template/ProForma_BLANK_master.xlsm)")
    ap.add_argument("--out", default=str(ROOT / "build" / "_underwriting"))
    ap.add_argument("--name", help="deal name (default: street address from DD)")
    ap.add_argument("--selftest", action="store_true", help="round-trip the example .xlsm and verify cells")
    a = ap.parse_args()

    if a.selftest:
        sys.exit(0 if selftest(a.dd) else 1)
    if not a.dd:
        ap.error("DD checklist workbook required")
    if not a.template:
        ap.error("no template: pass --template or set $SOLA_UW_TEMPLATE to the master pro-forma .xlsm")

    dd = read_dd(a.dd)
    paths, meta = export(dd, a.template, a.out, deal_name=a.name)
    print(f"product: {meta['product']}  (resource {meta['resource_mapped']})")
    for f in meta["flags"]:
        print(f"  FLAG: {f}")
    print("wrote:")
    for p in paths:
        print(f"  {p}")


if __name__ == "__main__":
    main()
